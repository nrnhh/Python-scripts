# -*- coding: utf-8 -*-
import oracledb
import pandas as pd
import os
import logging
import sys
import traceback
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

# --------------- Konfiqurasiya (Azərbaycanca Ay Adları üçün) ---------------
import locale
try:
    locale.setlocale(locale.LC_TIME, 'az_AZ.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'Azerbaijan_Azerbaijan.1251')
    except locale.Error:
        pass 
# -------------------------------------------------------------------------


# ------------------------- KONFİQURASİYA -------------------------
username = "NURAN"
password = "Nuran..2024!!"
dsn = "172.18.79.23:1521/FONETAZ"
oracle_client_path = r"C:\instant\instantclient_23_9" 

output_filename = "aylara_gore_100_faiz_endirim_hesabat.xlsx"
output_path = os.path.join(os.path.expanduser("~"), "Desktop", output_filename)

# ------------------------- SQL SORĞUSU (Final SELECT Yenidən Yazıldı) -------------------------
ENDIRIM_DETAIL_SQL_QUERY = r"""
WITH
SVC AS (
    SELECT
        X.HK_ID,
        MIN(TRUNC(X.HK_MURACAATTAR)) AS MURACIET_TARIXI,
        MIN(X.HK_KODU) AS KART_NOMRESI,
        MIN(X.HK_ADI || ' ' || X.HK_SOYADI) AS AD_SOYAD,
        SUM(NVL(T.HI_MIKTAR,0) * NVL(T.HI_HFIYAT,0)) AS XIDMET_QIYMETI,
        SUM(NVL(T.HI_MIKTAR,0) * NVL(T.HI_IFIYAT,0)) AS ENDIRIMSIZ_XIDMET_QIYMETI
    FROM FONETHBYS.H_HASTAKAYIT_ALT T
    JOIN FONETHBYS.H_HASTAKAYIT X ON X.HK_ID = T.HI_KAYITID
    WHERE X.HK_DURUM = '+'
        AND X.HK_HSINIFID IN (1)
        AND (
            NVL(T.HI_HIND_YUZDE,0) = 100
            OR (NVL(T.HI_HFIYAT,0) = 0 AND NVL(T.HI_IFIYAT,0) > 0)
        )
    GROUP BY X.HK_ID
),
DRG AS (
    SELECT
        X.HK_ID,
        SUM(NVL(Z.TUTAR,0)) AS DERMAN_QIYMETI,
        SUM(NVL(Z.IFIYAT,0) * NVL(Z.MIKTAR,0)) AS DERMAN_ENDIRIMSIZ_QIYMETI
    FROM FONETHBYS.V_STOK_LIST_HASTA_ILACSARF Z
    JOIN FONETHBYS.H_HASTAKAYIT X
        ON X.HK_ID = Z.HASTA_HKID
    WHERE X.HK_DURUM = '+'
        AND X.HK_HSINIFID IN (1)
        AND NVL(Z.TUTAR,0) = 0
        AND (NVL(Z.IFIYAT,0) * NVL(Z.MIKTAR,0)) > 0
        AND Z.PAKETDURUM='1'
    GROUP BY X.HK_ID
),
ALL_LOGS AS (
    SELECT Z.HI_KAYITID AS HK_ID, V.ACTION_DATE, V.ADSOYAD
    FROM FONETLOG.V_LOG_LIST V
    JOIN FONETHBYS.H_HASTAKAYIT_ALT Z ON Z.HI_ID=V.TABLE_PK_ID
    WHERE V.ACTION_DATE >= DATE '2025-01-01' AND V.ACTION_DATE < DATE '2026-01-01'
        AND V.USER_ID IN (49) AND V.OBJECT_ID IN (24702)
        AND V.COLUMN_NAME='HI_HIND_YUZDE' AND V.NEW_VALUE='100'
    UNION ALL
    SELECT S.HASTA_HKID AS HK_ID, V.ACTION_DATE, V.ADSOYAD
    FROM FONETLOG.V_LOG_LIST V
    JOIN FONETHBYS.STOK_CIKISDETAY S ON S.ID=V.TABLE_PK_ID
    WHERE V.ACTION_DATE >= DATE '2025-01-01' AND V.ACTION_DATE < DATE '2026-01-01'
        AND V.USER_ID IN (49) AND V.OBJECT_ID IN (24768)
        AND V.COLUMN_NAME='HASTA_INDYUZDE' AND V.NEW_VALUE='100'
),
MIN_LOG AS (
    SELECT
        HK_ID,
        MIN(ACTION_DATE) KEEP (DENSE_RANK FIRST ORDER BY ACTION_DATE ASC) AS ENDIRIM_TARIXI
    FROM ALL_LOGS
    GROUP BY HK_ID
)
SELECT
    Q.MURACIET_TARIXI,
    Q.KART_NOMRESI,
    Q.AD_SOYAD AS XESTE_AD_SOYAD,
    L.ENDIRIM_TARIXI,
    (NVL(S.ENDIRIMSIZ_XIDMET_QIYMETI,0) - NVL(S.XIDMET_QIYMETI,0)) AS XIDMET_ENDIRIMI,
    (NVL(D.DERMAN_ENDIRIMSIZ_QIYMETI,0) - NVL(D.DERMAN_QIYMETI,0)) AS DERMAN_ENDIRIMI,
    ( (NVL(S.ENDIRIMSIZ_XIDMET_QIYMETI,0) - NVL(S.XIDMET_QIYMETI,0))
      + (NVL(D.DERMAN_ENDIRIMSIZ_QIYMETI,0) - NVL(D.DERMAN_QIYMETI,0)) ) AS UMUMI_ENDIRIM,
    ( NVL(S.XIDMET_QIYMETI,0) + NVL(D.DERMAN_QIYMETI,0) ) AS TOPLAM_QIYMET,
    ( NVL(S.ENDIRIMSIZ_XIDMET_QIYMETI,0) + NVL(D.DERMAN_ENDIRIMSIZ_QIYMETI,0) ) AS ENDIRIMSIZ_TOPLAM
FROM SVC S
LEFT JOIN DRG D ON D.HK_ID = S.HK_ID
LEFT JOIN MIN_LOG L ON L.HK_ID = S.HK_ID
INNER JOIN (SELECT HK_ID, MURACIET_TARIXI, KART_NOMRESI, AD_SOYAD FROM SVC) Q ON Q.HK_ID = S.HK_ID
WHERE ( (NVL(S.ENDIRIMSIZ_XIDMET_QIYMETI,0) - NVL(S.XIDMET_QIYMETI,0))
      + (NVL(D.DERMAN_ENDIRIMSIZ_QIYMETI,0) - NVL(D.DERMAN_QIYMETI,0)) ) > 0
    AND (NVL(S.ENDIRIMSIZ_XIDMET_QIYMETI,0) + NVL(D.DERMAN_ENDIRIMSIZ_QIYMETI,0)) >= 10
ORDER BY L.ENDIRIM_TARIXI ASC
"""

# ------------------------- LOGGING -------------------------
logging.basicConfig(format='%(asctime)s | %(levelname)s | %(message)s', level=logging.INFO)
logger = logging.getLogger("endirim-rapor")

# ------------------------- SUTUN BAŞLIQLARININ XƏRİTƏSİ -------------------------
COLUMN_MAP = {
    'MURACIET_TARIXI': 'Müraciət Tarixi',
    'KART_NOMRESI': 'Kart Nömrəsi',
    'XESTE_AD_SOYAD': 'Xəstə Ad/Soyad',
    'ENDIRIM_TARIXI': 'Endirim Tarixi',
    'XIDMET_ENDIRIMI': 'Xidmət Endirimi (AZN)',
    'DERMAN_ENDIRIMI': 'Dərman Endirimi (AZN)',
    'UMUMI_ENDIRIM': 'Ümumi Endirim (AZN)',
    'TOPLAM_QIYMET': 'Ödənilən Məbləğ (AZN)',
    'ENDIRIMSIZ_TOPLAM': 'Endirimsiz Ümumi (AZN)'
}

# ------------------------- FUNKSİYALAR -------------------------
def initialize_oracle_client(client_path: str):
    if not os.path.exists(client_path):
        logger.error(f"❌ Oracle Instant Client yolu tapılmadı: {client_path}")
        sys.exit(1)
    oracledb.init_oracle_client(lib_dir=client_path)

def clean_sql_query(sql: str) -> str:
    """SQL sorğusunu təmizləyir və bütün ardıcıl boşluq simvollarını (o cümlədən görünməyənləri) 
       tək bir adi boşluqla əvəz edir."""
    # Bu daha təhlükəsiz təmizləmə metodudur
    cleaned = re.sub(r'\s+', ' ', sql).strip()
    return cleaned

def get_month_sheet_name(date_value):
    try:
        return date_value.strftime('%Y %B')
    except:
        return "Tarix Yoxdur" 

def create_notes_sheet(wb):
    ws = wb.create_sheet("Qeyd", 0)
    
    red_bold_font = Font(bold=True, color="FF0000")
    
    notes_text = [
        "Hesabatın Şərtləri:",
        "1. Hesabatda yalnız 'Ödənişli xəstələr' sinifinə (HK_HSINIFID=1) aid olan xəstələr nəzərə alınmışdır.",
        "2. Əlavə olaraq, ümumi endirim məbləği > 10 AZN şərti tətbiq edilmişdir. Bu yanaşma, 'Tibb bacısı' xidmətlərinin mövcudluğu səbəbindən real endirim məbləğinin düzgün müəyyən edilməsi üçün istifadə olunmuşdur.",
        "3. Nəticə etibarilə, ümumi endirim məbləği 10 AZN-dən az olan xəstələr hesabatda əks olunmamışdır.",
        "",
        "Məlumat: 'Endirim Tarixi' sütunu log məlumatlarından götürülmüşdür və 100% endirimin edildiyi ilk tarixi göstərir."
    ]
    
    ws.cell(row=1, column=1, value="Qeyd:").font = red_bold_font
    
    for r_idx, note in enumerate(notes_text, 2):
        cell = ws.cell(row=r_idx, column=1)
        cell.value = note 
        cell.font = red_bold_font
        
    ws.column_dimensions['A'].width = 120
    ws.row_dimensions[1].height = 20
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='top')

def write_sheet_data(ws, df_month: pd.DataFrame, sheet_title: str):
    
    col_header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    group_total_fill = PatternFill(start_color="FEE0A6", end_color="FEE0A6", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    internal_cols = list(COLUMN_MAP.keys()) # 9 sütun
    display_cols = list(COLUMN_MAP.values())
    
    start_row = 1
    
    for c_idx, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = bold_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    current_row = start_row + 1
    data_start_row = current_row
    
    for _, row in df_month.iterrows():
        for c_idx, col_key in enumerate(internal_cols, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=row[col_key])
            
            if col_key in ['MURACIET_TARIXI']:
                cell.number_format = 'yyyy-mm-dd'
            elif col_key in ['ENDIRIM_TARIXI']:
                cell.number_format = 'yyyy-mm-dd hh:mm'
            elif col_key in ['XIDMET_ENDIRIMI', 'DERMAN_ENDIRIMI', 'UMUMI_ENDIRIM', 'TOPLAM_QIYMET', 'ENDIRIMSIZ_TOPLAM']:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif col_key in ['KART_NOMRESI']:
                cell.alignment = right_align
            elif col_key in ['XESTE_AD_SOYAD']:
                cell.alignment = Alignment(horizontal='left')
            
            cell.border = thin_border
            
        current_row += 1
    
    data_end_row = current_row - 1
    num_cols = len(internal_cols) # 9
    
    # CƏM yazısı 4-cü sütuna düşür (Endirim Tarixi)
    ws.cell(row=current_row, column=4, value=f"{sheet_title} CƏMİ:").font = bold_font
    ws.cell(row=current_row, column=4).fill = group_total_fill
    ws.cell(row=current_row, column=4).alignment = right_align
    
    # Cəm düsturları 5-ci sütundan (Xidmət Endirimi) başlayır.
    for col_idx in range(5, num_cols + 1): 
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
        cell = ws.cell(row=current_row, column=col_idx, value=formula)
        cell.font = bold_font
        cell.number_format = '#,##0.00'
        cell.fill = group_total_fill
        cell.alignment = right_align
        
    for c in range(1, num_cols + 1):
        ws.cell(row=current_row, column=c).border = thin_border
    
    # Sütun enləri
    ws.column_dimensions['A'].width = 15 # Müraciət Tarixi
    ws.column_dimensions['B'].width = 15 # Kart Nömrəsi
    ws.column_dimensions['C'].width = 30 # Xəstə Ad/Soyad
    ws.column_dimensions['D'].width = 22 # Endirim Tarixi
    ws.column_dimensions['E'].width = 15 # Xidmət Endirimi
    ws.column_dimensions['F'].width = 15 # Dərman Endirimi
    ws.column_dimensions['G'].width = 18 # Ümumi Endirim
    ws.column_dimensions['H'].width = 18 # Ödənilən Məbləğ
    ws.column_dimensions['I'].width = 18 # Endirimsiz Ümumi
    
    ws.freeze_panes = 'A2' 

def generate_report():
    
    try:
        cleaned_sql = clean_sql_query(ENDIRIM_DETAIL_SQL_QUERY)
        initialize_oracle_client(oracle_client_path)

        with oracledb.connect(user=username, password=password, dsn=dsn) as conn:
            logger.info("Oracle verilənlər bazasından məlumat çəkilir...")
            df = pd.read_sql(cleaned_sql, con=conn)

        if df.empty:
            print("⚠️ Heç bir sətir tapılmadı. Fayl yaradılmadı.")
            return

        df.columns = [col.upper() for col in df.columns]
        
        df['AY_ACARI'] = df['ENDIRIM_TARIXI'].apply(
            lambda x: x.strftime('%Y-%m') if pd.notna(x) else 'N/A'
        )
        
        df_valid = df[df['AY_ACARI'] != 'N/A'].copy()

        # HK_ID artıq final SELECT-də yoxdur, sıralama yalnız Endirim Tarixi ilə qalır
        df_valid = df_valid.sort_values(by=['ENDIRIM_TARIXI'], ascending=[True])

        wb = Workbook()
        wb.remove(wb.active) 
        
        create_notes_sheet(wb)

        all_monthly_totals = []
        
        grouped = df_valid.groupby('AY_ACARI', sort=True)
        
        for ay_acari, group in grouped:
            
            month_name = get_month_sheet_name(group['ENDIRIM_TARIXI'].min())
            
            sheet_name = month_name.replace(' ', '_').replace('/', '_').replace(':', '_')[:31]
            
            ws = wb.create_sheet(sheet_name)
            
            write_sheet_data(ws, group, month_name)
            
            monthly_total = group[['UMUMI_ENDIRIM', 'TOPLAM_QIYMET', 'ENDIRIMSIZ_TOPLAM']].sum().to_dict()
            monthly_total['AY'] = month_name
            all_monthly_totals.append(monthly_total)


        if all_monthly_totals:
            total_df = pd.DataFrame(all_monthly_totals)
            ws_total = wb.create_sheet("Ümumi Cəm", 1) 
            
            grand_total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))


            header_cols = ["AY", "ÜMUMI ENDIRIM (AZN)", "TOPLAM QIYMƏT (AZN)", "ENDIRIMSIZ TOPLAM (AZN)"]
            for c_idx, col_name in enumerate(header_cols, 1):
                cell = ws_total.cell(row=1, column=c_idx, value=col_name)
                cell.font = bold_font
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            
            for r_idx, row in total_df.iterrows():
                current_data_row = r_idx + 2
                ws_total.cell(row=current_data_row, column=1, value=row['AY']).border = thin_border
                for c_idx, col in enumerate(['UMUMI_ENDIRIM', 'TOPLAM_QIYMET', 'ENDIRIMSIZ_TOPLAM'], 2):
                    cell = ws_total.cell(row=current_data_row, column=c_idx, value=row[col])
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align
                    cell.border = thin_border
            
            final_row = len(total_df) + 1
            grand_total_row = final_row + 1
            
            cell_text = ws_total.cell(row=grand_total_row, column=1, value="BÖYÜK CƏM")
            cell_text.font = bold_font
            cell_text.fill = grand_total_fill
            cell_text.border = thin_border
            
            for c_idx in range(2, 5):
                col_letter = get_column_letter(c_idx)
                formula = f"=SUM({col_letter}2:{col_letter}{final_row})"
                cell = ws_total.cell(row=grand_total_row, column=c_idx, value=formula)
                cell.font = bold_font
                cell.fill = grand_total_fill
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
                cell.border = thin_border
            
            ws_total.column_dimensions['A'].width = 25
            ws_total.column_dimensions['B'].width = 25
            ws_total.column_dimensions['C'].width = 25
            ws_total.column_dimensions['D'].width = 25


        wb.save(output_path)
        
        print(f"\n✅ Rapor uğurla yaradıldı və saxlanıldı: {output_path}")

    except oracledb.DatabaseError as e:
        logger.error(f"❌ Veritabanı xətası: {e}")
        print(f"❌ Veritabanı xətası: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"❌ Gözlənilməz Xəta: {e}", exc_info=True)
        print(f"❌ Xəta: {e}\n{traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    for m in ['oracledb', 'pandas', 'openpyxl', 're']:
        try:
            __import__(m)
        except ImportError:
            print(f"❌ '{m}' modulu quraşdırılmayıb. Quraşdırın: pip install {m}")
            sys.exit(1)
    
    generate_report()
