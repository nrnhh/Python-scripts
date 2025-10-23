# -*- coding: utf-8 -*-
import oracledb
import pandas as pd
import os
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
import sys
import traceback
import re
from itertools import groupby

# ------------------------- LOGGING -------------------------
logging.basicConfig(
    format='%(asctime)s | %(levelname)s | %(message)s',
    level=logging.INFO
)
logger = logging.getLogger("hekime-aitan-rapor")

# ------------------------- KONFİQURASİYA -------------------------
username = "NURAN"
password = "Nuran..2024!!"
dsn      = "172.18.79.23:1521/FONETAZ"
oracle_client_path = r"C:\instant\instantclient_23_9"

# Çıxış faylı
base_output_path = os.path.join(os.path.expanduser("~"), "Desktop", "raport.xlsx")

# Ay adı və rənglər
MONTH_ORDER = ['2025-05','2025-06','2025-07','2025-08','2025-09']
MONTH_NAME  = {'2025-05':'May','2025-06':'İyun','2025-07':'İyul','2025-08':'Avqust','2025-09':'Sentyabr'}
MONTH_COLOR = {'2025-05':"FFF0C2",'2025-06':"E0E0F5",'2025-07':"C8F0D8",'2025-08':"D9C2F5",'2025-09':"C2E0FF"}

# Xəstə sinifləri
CLASS_ORDER = [
    'Ödənişli Xəstələr','Sigorta Xəstələri','Dövlet Teşkilatları','Gönderen Hekim Teşkilatları',
    'Baku Medical Plaza (işçilər)','Savitar Group','BMP 100%','Korparativ','Bmp açılış günü','Sosial Tərəfdaş','Digər'
]

# Metriklər və başlıq mətnləri
METRICS = ['Xeste','Muayine','Lab']
METRIC_TEXT = {
    'XESTE':   'Xəstə sayı',
    'MUAYINE': 'Müayinə sayı',
    'LAB':     'Labaratoriya göndəriş sayı'
}

# ------------------------- SQL (HAMISI T.HK_KODU ÜZRƏ DISTINCT) -------------------------
SQL_QUERY = r"""
WITH
-- Hekim terefinden gosterilen xidmetler (Xeste sayi - DISTINCT HK_KODU)
eden AS (
  SELECT
      TO_CHAR(TRUNC(t.HI_TARIH,'MM'),'YYYY-MM') AS Ay,
      CASE
        WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^A-?AYAKTAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('AYAKTAN','A') THEN 'Ambulator'
        WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^Y-?YATAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('YATAN','Y') THEN 'Yatan'
        ELSE 'Digər'
      END AS Turu,
      t.HK_HSINIFID AS Sinif,
      CASE t.HK_HSINIFID
        WHEN 1 THEN 'Ödənişli Xəstələr'
        WHEN 2 THEN 'Sigorta Xəstələri'
        WHEN 5 THEN 'Dövlet Teşkilatları'
        WHEN 6 THEN 'Gönderen Hekim Teşkilatları'
        WHEN 8 THEN 'Baku Medical Plaza (işçilər)'
        WHEN 11 THEN 'Savitar Group'
        WHEN 12 THEN 'BMP 100%'
        WHEN 13 THEN 'Korparativ'
        WHEN 14 THEN 'Bmp açılış günü'
        WHEN 15 THEN 'Sosial Tərəfdaş'
        ELSE 'Digər'
      END AS Sinif_Adi,
      p1.P_KODU AS Hekim_Kodu,
      un1.pu_unvan AS Vezife,
      p1.P_AD || ' ' || p1.P_SOYAD AS Hekim_Adi,
      COUNT(DISTINCT t.HK_KODU) AS Xeste -- Bütün xidmətlər üçün fərqli xəstə sayı
  FROM fonethbys.V_IST_GENEL_HIZMET t
  JOIN fonethbysadm.H_PERSON p1 ON p1.P_ID = t.YPERSID
  JOIN fonethbysadm.H_PUNVAN un1 ON un1.PU_ID = p1.P_UNVANID
  WHERE t.HI_TARIH >= DATE '2025-05-01'
    AND t.HI_TARIH < DATE '2025-10-01'
  GROUP BY TO_CHAR(TRUNC(t.HI_TARIH,'MM'),'YYYY-MM'),
           CASE
             WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^A-?AYAKTAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('AYAKTAN','A') THEN 'Ambulator'
             WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^Y-?YATAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('YATAN','Y') THEN 'Yatan'
             ELSE 'Digər'
           END,
           t.HK_HSINIFID, p1.P_KODU, p1.P_AD, p1.P_SOYAD, un1.pu_unvan
),
-- Hekim terefinden gosterilen xidmetler (Muayine sayi - DISTINCT HK_KODU)
muayine AS (
  SELECT
      TO_CHAR(TRUNC(t.HI_TARIH,'MM'),'YYYY-MM') AS Ay,
      CASE
        WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^A-?AYAKTAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('AYAKTAN','A') THEN 'Ambulator'
        WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^Y-?YATAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('YATAN','Y') THEN 'Yatan'
        ELSE 'Digər'
      END AS Turu,
      t.HK_HSINIFID AS Sinif,
      CASE t.HK_HSINIFID
        WHEN 1 THEN 'Ödənişli Xəstələr'
        WHEN 2 THEN 'Sigorta Xəstələri'
        WHEN 5 THEN 'Dövlet Teşkilatları'
        WHEN 6 THEN 'Gönderen Hekim Teşkilatları'
        WHEN 8 THEN 'Baku Medical Plaza (işçilər)'
        WHEN 11 THEN 'Savitar Group'
        WHEN 12 THEN 'BMP 100%'
        WHEN 13 THEN 'Korparativ'
        WHEN 14 THEN 'Bmp açılış günü'
        WHEN 15 THEN 'Sosial Tərəfdaş'
        ELSE 'Digər'
      END AS Sinif_Adi,
      p1.P_KODU AS Hekim_Kodu,
      un1.pu_unvan AS Vezife,
      p1.P_AD || ' ' || p1.P_SOYAD AS Hekim_Adi,
      COUNT(DISTINCT t.HK_KODU) AS Muayine -- Fərqli Müayinə alan Xəstə Sayı (Düzəliş edildi)
  FROM fonethbys.V_IST_GENEL_HIZMET t
  JOIN fonethbysadm.H_PERSON p1 ON p1.P_ID = t.YPERSID
  JOIN fonethbysadm.H_PUNVAN un1 ON un1.PU_ID = p1.P_UNVANID
  WHERE t.HI_TARIH >= DATE '2025-05-01'
    AND t.HI_TARIH < DATE '2025-10-01'
    -- Yalniz Muayine kimi qeyd olunan xidmetler:
    AND (t.IS_RESMIKOD LIKE 'MUA%' AND t.IS_RESMIKOD NOT LIKE '%KONT'
         OR t.IS_RESMIKOD LIKE '%KONT')
  GROUP BY TO_CHAR(TRUNC(t.HI_TARIH,'MM'),'YYYY-MM'),
           CASE
             WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^A-?AYAKTAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('AYAKTAN','A') THEN 'Ambulator'
             WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^Y-?YATAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('YATAN','Y') THEN 'Yatan'
             ELSE 'Digər'
           END,
           t.HK_HSINIFID, p1.P_KODU, p1.P_AD, p1.P_SOYAD, un1.pu_unvan
),
-- Hekim terefinden gonderilen labaratoriyalar (Lab sayi - DISTINCT HK_KODU)
gonderen AS (
  SELECT
      TO_CHAR(TRUNC(t.HI_TARIH,'MM'),'YYYY-MM') AS Ay,
      CASE
        WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^A-?AYAKTAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('AYAKTAN','A') THEN 'Ambulator'
        WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^Y-?YATAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('YATAN','Y') THEN 'Yatan'
        ELSE 'Digər'
      END AS Turu,
      t.HK_HSINIFID AS Sinif,
      CASE t.HK_HSINIFID
        WHEN 1 THEN 'Ödənişli Xəstələr'
        WHEN 2 THEN 'Sigorta Xəstələri'
        WHEN 5 THEN 'Dövlet Teşkilatları'
        WHEN 6 THEN 'Gönderen Hekim Teşkilatları'
        WHEN 8 THEN 'Baku Medical Plaza (işçilər)'
        WHEN 11 THEN 'Savitar Group'
        WHEN 12 THEN 'BMP 100%'
        WHEN 13 THEN 'Korparativ'
        WHEN 14 THEN 'Bmp açılış günü'
        WHEN 15 THEN 'Sosial Tərəfdaş'
        ELSE 'Digər'
      END AS Sinif_Adi,
      p2.P_KODU AS Hekim_Kodu,
      un2.pu_unvan AS Vezife,
      p2.P_AD || ' ' || p2.P_SOYAD AS Hekim_Adi,
      COUNT(DISTINCT t.HK_KODU) AS Lab -- Fərqli Lab göndərişi alan Xəstə Sayı (Düzəliş edildi)
  FROM fonethbys.V_IST_GENEL_HIZMET t
  JOIN fonethbysadm.H_PERSON p2 ON p2.P_ID = t.IPERSID
  JOIN fonethbysadm.H_PUNVAN un2 ON un2.PU_ID = p2.P_UNVANID
  WHERE t.HI_TARIH >= DATE '2025-05-01'
    AND t.HI_TARIH < DATE '2025-10-01'
    AND t.ISLEMGRUPADI = 'Laboratuvar'
  GROUP BY TO_CHAR(TRUNC(t.HI_TARIH,'MM'),'YYYY-MM'),
           CASE
             WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^A-?AYAKTAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('AYAKTAN','A') THEN 'Ambulator'
             WHEN REGEXP_LIKE(UPPER(TRIM(t.HK_HASTATURU)), '^Y-?YATAN') OR UPPER(TRIM(t.HK_HASTATURU)) IN ('YATAN','Y') THEN 'Yatan'
             ELSE 'Digər'
           END,
           t.HK_HSINIFID, p2.P_KODU, p2.P_AD, p2.P_SOYAD, un2.pu_unvan
),
-- Neticelerin birleşdirilmesi
merged AS (
  SELECT Ay, Turu, Sinif, Sinif_Adi, Hekim_Kodu, Vezife, Hekim_Adi, Xeste, 0 AS Muayine, 0 AS Lab FROM eden
  UNION ALL
  SELECT Ay, Turu, Sinif, Sinif_Adi, Hekim_Kodu, Vezife, Hekim_Adi, 0 AS Xeste, Muayine, 0 AS Lab FROM muayine
  UNION ALL
  SELECT Ay, Turu, Sinif, Sinif_Adi, Hekim_Kodu, Vezife, Hekim_Adi, 0 AS Xeste, 0 AS Muayine, Lab FROM gonderen
)
-- Son netice (Bütün metriklərin cəmlənməsi)
SELECT
  Ay, Turu, Sinif, Sinif_Adi, Hekim_Kodu, Vezife, Hekim_Adi,
  SUM(Xeste) AS Xeste,
  SUM(Muayine) AS Muayine,
  SUM(Lab) AS Lab
FROM merged
GROUP BY Ay, Turu, Sinif, Sinif_Adi, Hekim_Kodu, Vezife, Hekim_Adi
ORDER BY Hekim_Adi, Turu, Ay, Sinif
"""

# ------------------------- YARDIMÇI FUNKSİYALAR -------------------------
def initialize_oracle_client(client_path: str):
    if not os.path.exists(client_path):
        print(f"❌ Oracle Instant Client yolu tapılmadı: {client_path}")
        sys.exit(1)
    oracledb.init_oracle_client(lib_dir=client_path)
    logger.info(f"Oracle Instant Client başladıldı: {client_path}")

def get_unique_output_path(base_path: str) -> str:
    if not os.path.exists(base_path):
        return base_path
    base, ext = os.path.splitext(base_path)
    i = 1
    while True:
        cand = f"{base}_{i}{ext}"
        if not os.path.exists(cand):
            return cand
        i += 1

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[\[\]\*?:/\\]', '_', name)
    return (name or "Sheet")[:31]

def get_scalar(val, default=None):
    try:
        if isinstance(val, pd.Series):
            return val.iloc[0] if not val.empty else default
        if hasattr(val, "item"):
            try:
                return val.item()
            except Exception:
                return val
        return val if val is not None else default
    except Exception:
        return val if val is not None else default

# ------------------------- EXCEL YAZIMI -------------------------
def write_sheet_with_multiheader(workbook: Workbook, sheet_name: str,
                                 wide_df: pd.DataFrame, ordered_cols: list):
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sanitize_sheet_name(sheet_name))

    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right', vertical='center')
    bold   = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

    title = f"{sheet_name} – Ay → (Xəstə sayı / Müayinə sayı / Labaratoriya göndəriş sayı) → Xəstə sinifi"
    total_cols = 4 + len(ordered_cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, total_cols))
    c = ws.cell(row=1, column=1, value=title); c.alignment = center; c.font = title_font
    c.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    base_headers = ['No','Kod','Vezife','Hekim Adı']
    for i, h in enumerate(base_headers, 1):
        ws.merge_cells(start_row=2, start_column=i, end_row=4, end_column=i)
        cell = ws.cell(row=2, column=i, value=h)
        cell.alignment = center; cell.font = bold
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        cell.border = thin

    col = 5
    for ay, group_by_ay in groupby(ordered_cols, key=lambda t: t[0]):
        group_by_ay = list(group_by_ay)
        ay_span = len(group_by_ay)
        if ay_span > 0:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + ay_span - 1)
            c1 = ws.cell(row=2, column=col, value=MONTH_NAME.get(ay, ay))
            c1.alignment = center; c1.font = bold
            c1.fill = PatternFill(start_color=MONTH_COLOR.get(ay, "FFFFFF"),
                                  end_color=MONTH_COLOR.get(ay, "FFFFFF"), fill_type="solid")
        for met, group_by_met in groupby(group_by_ay, key=lambda t: t[1]):
            group_by_met = list(group_by_met)
            metric_span = len(group_by_met)
            if metric_span > 0:
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + metric_span - 1)
                c2 = ws.cell(row=3, column=col, value=METRIC_TEXT.get(met.upper(), met))
                c2.alignment = center; c2.font = bold
                c2.fill = PatternFill(start_color=MONTH_COLOR.get(ay, "FFFFFF"),
                                      end_color=MONTH_COLOR.get(ay, "FFFFFF"), fill_type="solid")
            for _, _, sinif in group_by_met:
                c3 = ws.cell(row=4, column=col, value=sinif)
                c3.alignment = center; c3.font = bold
                c3.fill = PatternFill(start_color=MONTH_COLOR.get(ay, "FFFFFF"),
                                      end_color=MONTH_COLOR.get(ay, "FFFFFF"), fill_type="solid")
                col += 1

    start_row = 5
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30

    for ridx, r in wide_df.iterrows():
        rr = start_row + ridx
        no_val       = get_scalar(r.get('No', ridx+1), ridx+1)
        kod_val      = get_scalar(r.get('Kod', ''), '')
        vezife_val   = get_scalar(r.get('Vezife', ''), '')
        hekim_val    = get_scalar(r.get('Hekim Adı', ''), '')
        try: no_val = int(no_val)
        except: pass

        ws.cell(row=rr, column=1, value=no_val).alignment = center; ws.cell(row=rr, column=1).border = thin
        ws.cell(row=rr, column=2, value=str(kod_val)).alignment = center; ws.cell(row=rr, column=2).border = thin
        ws.cell(row=rr, column=3, value=str(vezife_val)).alignment = center; ws.cell(row=rr, column=3).border = thin
        ws.cell(row=rr, column=4, value=str(hekim_val)).alignment = left;    ws.cell(row=rr, column=4).border = thin

        ccol = 5
        for tup in ordered_cols:
            val = r.get(tup, 0); val = get_scalar(val, 0)
            try: val = int(val)
            except: pass
            ws.cell(row=rr, column=ccol, value=val).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=rr, column=ccol).border = thin
            ccol += 1

    max_row = start_row + len(wide_df) - 1
    max_col = 4 + len(ordered_cols)
    for rr in range(2, max_row + 1):
        for cc in range(1, max_col + 1):
            ws.cell(row=rr, column=cc).border = thin

    for cc in range(5, max_col + 1):
        ws.column_dimensions[get_column_letter(cc)].width = 12

    ws.freeze_panes = 'A5'

# ------------------------- HESABAT GENERATORU -------------------------
def generate_report():
    try:
        initialize_oracle_client(oracle_client_path)

        output_path = get_unique_output_path(base_output_path)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        logger.info("Veritabanına bağlanılır...")
        with oracledb.connect(user=username, password=password, dsn=dsn) as conn:
            logger.info("SQL işlədilir...")
            df = pd.read_sql(SQL_QUERY, con=conn)

        if df.empty:
            logger.warning("Sorğu boş nəticə qaytardı.")
            print("⚠️ Heç bir sətir tapılmadı.")
            return

        df = df.rename(columns=str.upper)

        months_found = [m for m in MONTH_ORDER if m in df['AY'].unique().tolist()]

        workbook = Workbook()
        workbook.remove(workbook.active)

        for turu, df_t in df.groupby('TURU'):
            wide = pd.pivot_table(
                df_t,
                index=['HEKIM_KODU','VEZIFE','HEKIM_ADI'],
                columns=['AY','SINIF_ADI'],
                values=['XESTE','MUAYINE','LAB'],
                aggfunc='sum',
                fill_value=0
            ).reorder_levels([1,0,2], axis=1).sort_index(axis=1)

            existing_cols = set(wide.columns.tolist())
            ordered_cols = []
            for ay in months_found:
                for met in METRICS:
                    for cls in CLASS_ORDER:
                        probe = (ay, met.upper(), cls)
                        if probe in existing_cols:
                            ordered_cols.append(probe)

            wide_out = wide.reset_index()
            wide_out.insert(0, 'No', range(1, len(wide_out)+1))
            wide_out = wide_out.rename(columns={'HEKIM_KODU':'Kod','VEZIFE':'Vezife','HEKIM_ADI':'Hekim Adı'})

            sheet_name = str(turu)  # Ambulator / Yatan
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
            write_sheet_with_multiheader(workbook, sheet_name, wide_out, ordered_cols)

        workbook.save(output_path)
        print(f"✅ Rapor saxlanıldı: {output_path}")

    except oracledb.DatabaseError as e:
        print(f"❌ Veritabanı xətası: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Xəta: {e}\n{traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    for m in ['oracledb','pandas','openpyxl']:
        try:
            __import__(m)
        except ImportError:
            print(f"❌ {m} modulu quraşdırılmayıb. Quraşdırın: pip install {m}")
            sys.exit(1)
    generate_report()
