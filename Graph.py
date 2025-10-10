import oracledb
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import CommandHandler, MessageHandler, filters, ApplicationBuilder, ContextTypes, ConversationHandler, CallbackQueryHandler
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# Loglama qurmaq
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# Sabitlər və konfiqurasiya
custom_colors = ['red', 'blue', 'green', 'black', 'purple', 'orange', 'darkblue', 'darkgreen', 'brown', 'crimson', 'navy', 'darkred']

chat_id = '-1003177573983'
bot_token = "8386936045:AAE18mrqRK9SxgDjj0nbglv5B5sxH1Waaxo"
PASSIVE_DOCTOR_LINK = "https://t.me/+pVPrA77Eq08xOWZi"

# oracledb.init_oracle_client(lib_dir="/opt/oracle/instantclient_23_6")
oracledb.init_oracle_client(lib_dir=r"C:\instant\instantclient_23_9")
username = "NURAN"
password = "Nuran..2024!!"
dsn = "172.18.79.23:1521/FONETAZ"

# ConversationHandler vəziyyətləri
MAIN_MENU, ASK_DOCTOR, SELECT_DATE_RANGE, CONTINUE_DOCTOR, TOP_10_DATE, EXPORT_PROMPT, SELECT_SPECIALTY, SELECT_STAT_TYPE = range(8)

# SQL sorğuları (Dəyişilməyib)
query_template_specific = """
SELECT * FROM (
    SELECT
        '{tip}' AS TIP,
        T.{persid_col} AS HEKIM_ID,
        MAX(P.P_AD || ' ' || P.P_SOYAD) AS HEKIM_ADI,
        TO_CHAR(T.HI_TARIH, 'YYYY-MM') AS AY,
        EXTRACT(YEAR FROM T.HI_TARIH) AS YEAR_,
        EXTRACT(MONTH FROM T.HI_TARIH) AS MONTH_,
        COUNT(DISTINCT CASE WHEN T.HK_HASTATURU LIKE '%Ayaktan%' THEN T.HK_KODU END) AS AYAKTAN,
        COUNT(DISTINCT CASE WHEN T.HK_HASTATURU LIKE '%Yatan%' THEN T.HK_KODU END) AS YATAN
    FROM FONETHBYS.V_IST_GENEL_HIZMET T
    LEFT JOIN FONETHBYSADM.H_PERSON P ON T.{persid_col} = P.P_ID
    WHERE T.{persid_col} = (SELECT P_ID FROM FONETHBYSADM.H_PERSON WHERE P_KODU = :kod)
        AND T.HI_TARIH BETWEEN TO_DATE(:start_date,'DD.MM.YYYY') AND TO_DATE(:end_date,'DD.MM.YYYY')
    GROUP BY T.{persid_col}, TO_CHAR(T.HI_TARIH, 'YYYY-MM'), EXTRACT(YEAR FROM T.HI_TARIH), EXTRACT(MONTH FROM T.HI_TARIH)
)
ORDER BY YEAR_, MONTH_
"""

query_template_lab_specific = """
SELECT * FROM (
    SELECT
        '{tip}' AS TIP,
        T.{persid_col} AS HEKIM_ID,
        MAX(P.P_AD || ' ' || P.P_SOYAD) AS HEKIM_ADI,
        TO_CHAR(T.HI_TARIH, 'YYYY-MM') AS AY,
        EXTRACT(YEAR FROM T.HI_TARIH) AS YEAR_,
        EXTRACT(MONTH FROM T.HI_TARIH) AS MONTH_,
        COUNT(DISTINCT CASE WHEN T.HK_HASTATURU LIKE '%Ayaktan%' THEN T.HK_KODU END) AS AYAKTAN,
        COUNT(DISTINCT CASE WHEN T.HK_HASTATURU LIKE '%Yatan%' THEN T.HK_KODU END) AS YATAN
    FROM FONETHBYS.V_IST_GENEL_HIZMET T
    LEFT JOIN FONETHBYSADM.H_PERSON P ON T.{persid_col} = P.P_ID
    WHERE T.{persid_col} = (SELECT P_ID FROM FONETHBYSADM.H_PERSON WHERE P_KODU = :kod)
        AND LOWER(T.ISLEMGRUPADI) LIKE '%lab%'
        AND T.HI_TARIH BETWEEN TO_DATE(:start_date,'DD.MM.YYYY') AND TO_DATE(:end_date,'DD.MM.YYYY')
    GROUP BY T.{persid_col}, TO_CHAR(T.HI_TARIH, 'YYYY-MM'), EXTRACT(YEAR FROM T.HI_TARIH), EXTRACT(MONTH FROM T.HI_TARIH)
)
ORDER BY YEAR_, MONTH_
"""

query_template_specialty = """
SELECT
    TO_CHAR(A.TARIH, 'Month YYYY', 'NLS_DATE_LANGUAGE = AZERBAIJANI') AS AY_ADI,
    A.DOKTOR_ADI,
    A.HK_HASTATURU,
    A.ISLEMGRUPADI,
    A.UNVAN,
    SUM(A.EDEN_SAYISI) AS EDEN_SAYISI,
    SUM(A.GONDEREN_SAYISI) AS GONDEREN_SAYISI
FROM (
    -- Eden (xidməti verən) həkimlər
    SELECT
        TRUNC(T.HI_TARIH, 'MM') AS TARIH,
        YP.P_AD || ' ' || YP.P_SOYAD AS DOKTOR_ADI,
        T.HK_HASTATURU,
        T.ISLEMGRUPADI,
        PU.PU_UNVAN AS UNVAN,
        COUNT(DISTINCT T.HK_ID) AS EDEN_SAYISI,
        0 AS GONDEREN_SAYISI
    FROM FONETHBYS.V_IST_GENEL_HIZMET T
    LEFT JOIN FONETHBYSADM.H_PERSON YP ON T.YPERSID = YP.P_ID
    LEFT JOIN FONETHBYSADM.H_PUNVAN PU ON YP.P_UNVANID = PU.PU_ID
    WHERE T.HI_TARIH BETWEEN TO_DATE(:start_date, 'DD.MM.YYYY') AND TO_DATE(:end_date, 'DD.MM.YYYY')
        AND YP.P_UNVANID = :pu_id
    GROUP BY TRUNC(T.HI_TARIH, 'MM'), YP.P_AD, YP.P_SOYAD, T.HK_HASTATURU, T.ISLEMGRUPADI, PU.PU_UNVAN
    UNION ALL
    -- Göndərən həkimlər
    SELECT
        TRUNC(T.HI_TARIH, 'MM') AS TARIH,
        IP.P_AD || ' ' || IP.P_SOYAD AS DOKTOR_ADI,
        T.HK_HASTATURU,
        T.ISLEMGRUPADI,
        PU.PU_UNVAN AS UNVAN,
        0 AS EDEN_SAYISI,
        COUNT(DISTINCT T.HK_ID) AS GONDEREN_SAYISI
    FROM FONETHBYS.V_IST_GENEL_HIZMET T
    LEFT JOIN FONETHBYSADM.H_PERSON IP ON T.IPERSID = IP.P_ID
    LEFT JOIN FONETHBYSADM.H_PUNVAN PU ON IP.P_UNVANID = PU.PU_ID
    WHERE T.HI_TARIH BETWEEN TO_DATE(:start_date, 'DD.MM.YYYY') AND TO_DATE(:end_date, 'DD.MM.YYYY')
        AND IP.P_UNVANID = :pu_id
    GROUP BY TRUNC(T.HI_TARIH, 'MM'), IP.P_AD, IP.P_SOYAD, T.HK_HASTATURU, T.ISLEMGRUPADI, PU.PU_UNVAN
) A
GROUP BY A.TARIH, A.DOKTOR_ADI, A.HK_HASTATURU, A.ISLEMGRUPADI, A.UNVAN
ORDER BY A.TARIH, A.DOKTOR_ADI, A.HK_HASTATURU, A.ISLEMGRUPADI
"""

top_10_queries = {
    "eden_ayaktan": """
        SELECT * FROM (
            SELECT
                MAX(P.P_AD || ' ' || P.P_SOYAD) AS HEKIM_ADI,
                P.P_KODU,
                COUNT(DISTINCT T.HK_KODU) AS AYAKTAN,
                0 AS YATAN,
                COUNT(DISTINCT T.HK_KODU) AS CEMI
            FROM FONETHBYS.V_IST_GENEL_HIZMET T
            LEFT JOIN FONETHBYSADM.H_PERSON P ON T.YPERSID = P.P_ID
            WHERE T.HK_HASTATURU LIKE '%Ayaktan%'
                AND T.HI_TARIH BETWEEN TO_DATE(:start_date, 'DD.MM.YYYY') AND TO_DATE(:end_date, 'DD.MM.YYYY')
                AND P.P_UNVANID = :pu_id
            GROUP BY P.P_KODU
            ORDER BY CEMI DESC
        ) WHERE ROWNUM <= 10
    """,
    "eden_yatan": """
        SELECT * FROM (
            SELECT
                MAX(P.P_AD || ' ' || P.P_SOYAD) AS HEKIM_ADI,
                P.P_KODU,
                0 AS AYAKTAN,
                COUNT(DISTINCT T.HK_KODU) AS YATAN,
                COUNT(DISTINCT T.HK_KODU) AS CEMI
            FROM FONETHBYS.V_IST_GENEL_HIZMET T
            LEFT JOIN FONETHBYSADM.H_PERSON P ON T.YPERSID = P.P_ID
            WHERE T.HK_HASTATURU LIKE '%Yatan%'
                AND T.HI_TARIH BETWEEN TO_DATE(:start_date, 'DD.MM.YYYY') AND TO_DATE(:end_date, 'DD.MM.YYYY')
                AND P.P_UNVANID = :pu_id
            GROUP BY P.P_KODU
            ORDER BY CEMI DESC
        ) WHERE ROWNUM <= 10
    """,
    "gonderen_ayaktan": """
        SELECT * FROM (
            SELECT
                MAX(P.P_AD || ' ' || P.P_SOYAD) AS HEKIM_ADI,
                P.P_KODU,
                COUNT(DISTINCT T.HK_KODU) AS AYAKTAN,
                0 AS YATAN,
                COUNT(DISTINCT T.HK_KODU) AS CEMI
            FROM FONETHBYS.V_IST_GENEL_HIZMET T
            LEFT JOIN FONETHBYSADM.H_PERSON P ON T.IPERSID = P.P_ID
            WHERE T.HK_HASTATURU LIKE '%Ayaktan%'
                AND T.HI_TARIH BETWEEN TO_DATE(:start_date, 'DD.MM.YYYY') AND TO_DATE(:end_date, 'DD.MM.YYYY')
                AND P.P_UNVANID = :pu_id
            GROUP BY P.P_KODU
            ORDER BY CEMI DESC
        ) WHERE ROWNUM <= 10
    """,
    "gonderen_yatan": """
        SELECT * FROM (
            SELECT
                MAX(P.P_AD || ' ' || P.P_SOYAD) AS HEKIM_ADI,
                P.P_KODU,
                0 AS AYAKTAN,
                COUNT(DISTINCT T.HK_KODU) AS YATAN,
                COUNT(DISTINCT T.HK_KODU) AS CEMI
            FROM FONETHBYS.V_IST_GENEL_HIZMET T
            LEFT JOIN FONETHBYSADM.H_PERSON P ON T.IPERSID = P.P_ID
            WHERE T.HK_HASTATURU LIKE '%Yatan%'
                AND T.HI_TARIH BETWEEN TO_DATE(:start_date, 'DD.MM.YYYY') AND TO_DATE(:end_date, 'DD.MM.YYYY')
                AND P.P_UNVANID = :pu_id
            GROUP BY P.P_KODU
            ORDER BY CEMI DESC
        ) WHERE ROWNUM <= 10
    """,
    "lab_gonderen_ayaktan": """
        SELECT * FROM (
            SELECT
                MAX(P.P_AD || ' ' || P.P_SOYAD) AS HEKIM_ADI,
                P.P_KODU,
                COUNT(DISTINCT T.HK_KODU) AS AYAKTAN,
                0 AS YATAN,
                COUNT(DISTINCT T.HK_KODU) AS CEMI
            FROM FONETHBYS.V_IST_GENEL_HIZMET T
            LEFT JOIN FONETHBYSADM.H_PERSON P ON T.IPERSID = P.P_ID
            WHERE LOWER(T.ISLEMGRUPADI) LIKE '%lab%'
                AND T.HK_HASTATURU LIKE '%Ayaktan%'
                AND T.HI_TARIH BETWEEN TO_DATE(:start_date, 'DD.MM.YYYY') AND TO_DATE(:end_date, 'DD.MM.YYYY')
                AND P.P_UNVANID = :pu_id
            GROUP BY P.P_KODU
            ORDER BY CEMI DESC
        ) WHERE ROWNUM <= 10
    """,
    "lab_gonderen_yatan": """
        SELECT * FROM (
            SELECT
                MAX(P.P_AD || ' ' || P.P_SOYAD) AS HEKIM_ADI,
                P.P_KODU,
                0 AS AYAKTAN,
                COUNT(DISTINCT T.HK_KODU) AS YATAN,
                COUNT(DISTINCT T.HK_KODU) AS CEMI
            FROM FONETHBYS.V_IST_GENEL_HIZMET T
            LEFT JOIN FONETHBYSADM.H_PERSON P ON T.IPERSID = P.P_ID
            WHERE LOWER(T.ISLEMGRUPADI) LIKE '%lab%'
                AND T.HK_HASTATURU LIKE '%Yatan%'
                AND T.HI_TARIH BETWEEN TO_DATE(:start_date, 'DD.MM.YYYY') AND TO_DATE(:end_date, 'DD.MM.YYYY')
                AND P.P_UNVANID = :pu_id
            GROUP BY P.P_KODU
            ORDER BY CEMI DESC
        ) WHERE ROWNUM <= 10
    """
}

# Menyu düymələri (Dəyişilməyib)
def get_main_menu():
    keyboard = [
        [InlineKeyboardButton("📈 Statistik Qrafiklər (Fərdi)", callback_data='menu_stats')],
        [InlineKeyboardButton("📅 Tarixə görə Analiz (Qayıt)", callback_data='menu_date')],
        [
            InlineKeyboardButton("🏆 TOP 10 Eden Ayaktan", callback_data='menu_top10_eden_ayaktan'),
            InlineKeyboardButton("🏆 TOP 10 Eden Yatan", callback_data='menu_top10_eden_yatan')
        ],
        [
            InlineKeyboardButton("🏆 TOP 10 Gonderen Ayaktan", callback_data='menu_top10_gonderen_ayaktan'),
            InlineKeyboardButton("🏆 TOP 10 Gonderen Yatan", callback_data='menu_top10_gonderen_yatan')
        ],
        [
            InlineKeyboardButton("🏆 TOP 10 Lab Gonderen Ayaktan", callback_data='menu_top10_lab_ayaktan'),
            InlineKeyboardButton("🏆 TOP 10 Lab Gonderen Yatan", callback_data='menu_top10_lab_yatan')
        ],
        [InlineKeyboardButton("👨‍⚕️ Peşəyə görə Həkimlər", callback_data='menu_specialty')],
        [InlineKeyboardButton("❌ Passiv Həkimlər", callback_data='menu_passive')], 
        [InlineKeyboardButton("📊 Excel İxrac", callback_data='menu_excel')],
        [InlineKeyboardButton("🔁 Yeni Seçim", callback_data='menu_reset')]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_persistent_keyboard():
    keyboard = [[KeyboardButton("/start")]]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_date_range_buttons():
    keyboard = [
        [InlineKeyboardButton("3 Aylıq", callback_data='date_3m')],
        [InlineKeyboardButton("6 Aylıq", callback_data='date_6m')],
        [InlineKeyboardButton("1 İllik", callback_data='date_1y')],
        [InlineKeyboardButton("Ana menyuya qayıt", callback_data='return_main_menu')]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_continue_buttons(last_action):
    keyboard = [
        [
            InlineKeyboardButton("Başqa həkim əlavə et (Hə)", callback_data='continue_yes'),
            InlineKeyboardButton("Bitir (Yox)", callback_data='continue_no')
        ],
        # Yalnız fərdi axtarış (manual) rejimində tarix dəyişməyə imkan veririk
        *([
            [InlineKeyboardButton("📅 Tarix Aralığını Dəyiş", callback_data='back_to_date_range_manual')]
        ] if last_action == 'manual' else []),
        [InlineKeyboardButton("Ana menyuya qayıt", callback_data='return_main_menu')]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_export_prompt_buttons(last_action):
    keyboard = [
        [
            InlineKeyboardButton("📊 Excel olaraq göndər (Hə)", callback_data='export_yes'),
            InlineKeyboardButton("Ana menyuya qayıt (Yox)", callback_data='return_main_menu')
        ],
        # Yalnız fərdi axtarış (manual) rejimində tarix dəyişməyə imkan veririk
        *([
            [InlineKeyboardButton("📅 Tarix Aralığını Dəyiş", callback_data='back_to_date_range_manual')]
        ] if last_action == 'manual' else [])
    ]
    return InlineKeyboardMarkup(keyboard)

def get_specialty_buttons(specialties):
    keyboard = []
    for i in range(0, len(specialties), 2):
        row = [InlineKeyboardButton(specialties[i]['PU_UNVAN'], callback_data=f'specialty_{specialties[i]["PU_ID"]}')]
        if i + 1 < len(specialties):
            row.append(InlineKeyboardButton(specialties[i + 1]['PU_UNVAN'], callback_data=f'specialty_{specialties[i + 1]["PU_ID"]}'))
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("Ana menyuya qayıt", callback_data='return_main_menu')])
    return InlineKeyboardMarkup(keyboard)

def get_stat_type_buttons():
    # Düzəliş edilmiş funksiya
    keyboard = [
        [InlineKeyboardButton("Umumi Statistika", callback_data='stat_umumi')],
        [
            InlineKeyboardButton("Eden Ayaktan", callback_data='stat_eden_ayaktan'),
            InlineKeyboardButton("Eden Yatan", callback_data='stat_eden_yatan')
        ],
        [
            InlineKeyboardButton("Gonderen Ayaktan", callback_data='stat_gonderen_ayaktan'),
            InlineKeyboardButton("Gonderen Yatan", callback_data='stat_gonderen_yatan')
        ],
        [
            InlineKeyboardButton("Lab Gonderen Ayaktan", callback_data='stat_lab_gonderen_ayaktan'),
            InlineKeyboardButton("Lab Gonderen Yatan", callback_data='stat_lab_gonderen_yatan')
        ],
        [
            InlineKeyboardButton("Tarix aralığını dəyiş", callback_data='back_to_date_range'),
            InlineKeyboardButton("Peşə dəyiş", callback_data='back_to_specialty')
        ],
        [InlineKeyboardButton("Ana menyuya qayıt", callback_data='return_main_menu')]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_return_main_menu_button():
    keyboard = [[InlineKeyboardButton("Ana menyuya qayıt", callback_data='return_main_menu')]]
    return InlineKeyboardMarkup(keyboard)

# Start funksiyası (Dəyişilməyib)
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info(f"Received command or message from chat_id: {update.effective_chat.id}, message_id: {update.message.message_id}, text: {update.message.text}")
    context.user_data.clear()
    context.user_data['kodlar'] = []
    context.user_data['adlar'] = []
    try:
        await update.effective_chat.send_message(
            "Salam! 👨‍⚕️ Həkim statistikası botuna xoş gəldiniz!\n"
            "Zəhmət olmasa, aşağıdakı menyudan bir seçim edin.",
            reply_markup=get_main_menu(),
            reply_to_message_id=update.message.message_id
        )
        await update.effective_chat.send_message(
            "Botu yenidən başlatmaq üçün /start istifadə edin.",
            reply_markup=get_persistent_keyboard(),
            reply_to_message_id=update.message.message_id
        )
        logger.info("Bot başlatıldı, ana menyu göstərildi.")
        return MAIN_MENU
    except Exception as e:
        logger.error(f"Error in start function: {e}")
        await update.effective_chat.send_message(
            f"Xəta baş verdi: {e}\nYenidən cəhd edin və ya /start istifadə edin.",
            reply_markup=get_main_menu(),
            reply_to_message_id=update.message.message_id
        )
        return MAIN_MENU

# TOP 10 həkim funksiyası (Dəyişilməyib)
async def top_10_doctors(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("top_10_doctors funksiyası çağırıldı.")

    category = context.user_data.get('top10_category', 'eden_ayaktan')
    pu_id = context.user_data.get('specialty_id')
    specialty_name = context.user_data.get('specialty_name', 'Peşə')
    start_date = context.user_data['start_date']
    end_date = context.user_data['end_date']

    category_names = {
        'eden_ayaktan': 'Eden Ayaktan',
        'eden_yatan': 'Eden Yatan',
        'gonderen_ayaktan': 'Gonderen Ayaktan',
        'gonderen_yatan': 'Gonderen Yatan',
        'lab_gonderen_ayaktan': 'Lab Gonderen Ayaktan',
        'lab_gonderen_yatan': 'Lab Gonderen Yatan'
    }
    category_name = category_names[category]

    metric_config = {
        'eden_ayaktan': 'AYAKTAN',
        'eden_yatan': 'YATAN',
        'gonderen_ayaktan': 'AYAKTAN',
        'gonderen_yatan': 'YATAN',
        'lab_gonderen_ayaktan': 'AYAKTAN',
        'lab_gonderen_yatan': 'YATAN'
    }
    display_metric = metric_config[category]
    metric_label = 'Ayaktan' if display_metric == 'AYAKTAN' else 'Yatan'

    logger.info(f"Tarix aralığı: {start_date} - {end_date}, Peşə: {specialty_name} (PU_ID: {pu_id})")

    try:
        logger.info("Verilənlər bazasına qoşulur...")
        with oracledb.connect(user=username, password=password, dsn=dsn) as conn:
            logger.info(f"TOP 10 {category_name} sorğusu icra olunur...")
            top_10_df = pd.read_sql(
                top_10_queries[category],
                con=conn,
                params={"start_date": start_date, "end_date": end_date, "pu_id": int(pu_id)}
            )

            if top_10_df.empty:
                logger.warning(f"{start_date} - {end_date} tarixləri üçün {specialty_name} peşəsində {category_name} məlumat tapılmadı.")
                await update.effective_chat.send_message(
                    f"⚠️ {specialty_name} peşəsi üçün {start_date} - {end_date} tarixləri arasında {category_name} məlumat tapılmadı.\n"
                    "Zəhmət olmasa, başqa bir peşə və ya tarix aralığı seçin.",
                    reply_markup=get_return_main_menu_button(),
                    reply_to_message_id=update.effective_message.message_id
                )
                return MAIN_MENU

            logger.info(f"TOP 10 DataFrame: {top_10_df[['P_KODU', 'HEKIM_ADI']].to_dict()}")
            top_10_df = top_10_df[top_10_df['P_KODU'].notnull() & (top_10_df['P_KODU'] != '')]

            if top_10_df.empty:
                logger.warning(f"{start_date} - {end_date} tarixləri üçün {specialty_name} peşəsində {category_name} etibarlı P_KODU tapılmadı.")
                await update.effective_chat.send_message(
                    f"⚠️ {specialty_name} peşəsi üçün {start_date} - {end_date} tarixləri arasında {category_name} etibarlı həkim kodu tapılmadı.\n"
                    "Zəhmət olmasa, başqa bir peşə və ya tarix aralığı seçin.",
                    reply_markup=get_return_main_menu_button(),
                    reply_to_message_id=update.effective_message.message_id
                )
                return MAIN_MENU

            msg = f"🏆 TOP 10 {category_name} ({specialty_name}, {start_date} - {end_date})\n\n"
            for idx, row in top_10_df.iterrows():
                msg += f"{idx + 1}. {row['HEKIM_ADI']}\n"
                msg += f"    {metric_label}: {int(row[display_metric])}\n"
                msg += f"    Cəmi: {int(row['CEMI'])}\n\n"

            context.user_data['kodlar'] = top_10_df['P_KODU'].tolist()
            context.user_data['adlar'] = top_10_df['HEKIM_ADI'].tolist()
            context.user_data['last_action'] = 'top10'
            logger.info(f"TOP 10 {category_name} həkimləri saxlanıldı: {context.user_data['adlar']}")

            await update.effective_chat.send_message(
                msg,
                reply_to_message_id=update.effective_message.message_id
            )
            logger.info(f"TOP 10 {category_name} siyahısı göndərildi.")

            await update.effective_chat.send_message(
                f"📊 TOP 10 {category_name} həkimlər üçün qrafik hazırlanır, zəhmət olmasa gözləyin...",
                reply_to_message_id=update.effective_message.message_id
            )
            logger.info(f"{category_name} qrafiki yaradılır...")
            await generate_graph(update, context)

            await update.effective_chat.send_message(
                f"📊 TOP 10 {category_name} statistikasını Excel olaraq göndərilsin mi?\n"
                "Hə: Excel faylı göndəriləcək.\nYox: Ana menyuya qayıdacaqsınız.",
                reply_markup=get_export_prompt_buttons(context.user_data['last_action']),
                reply_to_message_id=update.effective_message.message_id
            )
            return EXPORT_PROMPT

    except Exception as e:
        logger.error(f"top_10_doctors xətası ({category_name}): {e}")
        await update.effective_chat.send_message(
            f"Xəta baş verdi: {e}\nZəhmət olmasa, ana menyudan yeni bir seçim edin.",
            reply_markup=get_return_main_menu_button(),
            reply_to_message_id=update.effective_message.message_id
        )
        return MAIN_MENU

# Excel ixrac funksiyası (ƏSAS DƏYİŞİKLİKLƏR)
async def export_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("export_to_excel funksiyası çağırıldı.")

    start_date = context.user_data.get('start_date', (datetime.now() - timedelta(days=180)).strftime('%d.%m.%Y'))
    end_date = context.user_data.get('end_date', datetime.now().strftime('%d.%m.%Y'))
    category = context.user_data.get('top10_category', 'hekim')
    specialty_name = context.user_data.get('specialty_name', 'Peşə')
    last_action = context.user_data.get('last_action')
    category_names = {
        'eden_ayaktan': 'Eden Ayaktan',
        'eden_yatan': 'Eden Yatan',
        'gonderen_ayaktan': 'Göndərən Ayaktan',
        'gonderen_yatan': 'Göndərən Yatan',
        'lab_gonderen_ayaktan': 'Laboratoriya Göndərən Ayaktan',
        'lab_gonderen_yatan': 'Laboratoriya Göndərən Yatan',
        'hekim': 'Həkim',
        'specialty': specialty_name
    }
    
    # Başlıq üçün kateqoriya adı
    if last_action == 'top10':
        category_name = f"TOP 10 {category_names.get(category, 'Həkim')}"
    elif last_action == 'specialty':
        category_name = f"Peşə ({specialty_name}) statistikası"
    else:
        category_name = 'Fərdi Həkim'

    try:
        if last_action == 'specialty' and 'specialty_df' in context.user_data:
            # Peşə statistikası üçün sadə DataFrame ixracı
            df = context.user_data['specialty_df'].copy()
            df.rename(columns={'AY_ADI': 'Ay / Peşə', 'DOKTOR_ADI': 'Həkim Adı', 'HK_HASTATURU': 'Xidmət Növü', 
                               'ISLEMGRUPADI': 'Qrup Adı', 'UNVAN': 'Peşə', 
                               'EDEN_SAYISI': 'Eden (Xidmət Sayı)', 'GONDEREN_SAYISI': 'Göndərən Sayı'}, inplace=True)
            df = df[['Ay / Peşə', 'Həkim Adı', 'Xidmət Növü', 'Qrup Adı', 'Peşə', 'Eden (Xidmət Sayı)', 'Göndərən Sayı']].fillna(0)

            # Peşə statistikasını ayrıca Excel faylı kimi göndər
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Peşə Statistikası', index=False)
            
            buffer.seek(0)
            file_name = f"statistikasi_{category}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=buffer,
                filename=file_name,
                caption=f"📊 {start_date} - {end_date} tarixləri üçün {category_name} statistikası",
                reply_to_message_id=update.effective_message.message_id
            )
            
            await update.effective_chat.send_message(
                "✅ Excel faylı göndərildi!\nNövbəti addım: Ana menyudan yeni bir seçim edə bilərsiniz.",
                reply_markup=get_return_main_menu_button(),
                reply_to_message_id=update.effective_message.message_id
            )
            return MAIN_MENU
        
        # Fərdi və TOP 10 üçün mürəkkəb formatlama (Ayrı Sheetlər və Pivot)
        else:
            if not context.user_data.get('kodlar'):
                raise ValueError("Həkim seçilməyib.")
            
            with oracledb.connect(user=username, password=password, dsn=dsn) as conn:
                query_sources = [
                    ("Eden-isci", "YPERSID", query_template_specific),
                    ("Gonderen-isci", "IPERSID", query_template_specific),
                    ("Gonderen-LAB", "IPERSID", query_template_lab_specific)
                ]

                all_data = []
                for kod, ad in zip(context.user_data['kodlar'], context.user_data['adlar']):
                    if kod is None or not isinstance(kod, str) or not kod.strip(): continue
                    for tip, persid_col, template in query_sources:
                        query = template.format(tip=tip, persid_col=persid_col)
                        df_temp = pd.read_sql(query, con=conn, params={"kod": kod.strip(), "start_date": start_date, "end_date": end_date})
                        if not df_temp.empty:
                            df_temp['HEKIM_ADI'] = ad # Həkim Adını əlavə et
                            all_data.append(df_temp)

                if not all_data:
                    raise ValueError("Seçilmiş həkimlər üçün məlumat tapılmadı.")

                df_raw = pd.concat(all_data, ignore_index=True)
                df_raw.rename(columns={'TIP': 'Məlumat Tipi', 'HEKIM_ADI': 'Həkim Adı', 'AY': 'Ay', 
                                       'AYAKTAN': 'Ayaktan Sayı', 'YATAN': 'Yatan Sayı', 
                                       'YEAR_': 'İl', 'MONTH_': 'Ay_Reqem'}, inplace=True)
                
                
                # --- YATAN VƏ AYAKTAN MƏLUMATLARININ AYRI QURUPLAŞDIRILMASI VƏ MƏLUMATIN TƏKRARLANMASI ---
                all_qroups = []
                
                # 1. AYAKTAN məlumatları olan sətirləri ayır (Ayaktan Sayı > 0 olan sətirlər)
                df_ayaktan = df_raw[df_raw['Ayaktan Sayı'] > 0].copy()
                if not df_ayaktan.empty:
                    df_ayaktan['Qrup_Adı'] = df_ayaktan['Məlumat Tipi'].str.replace('-isci', '').str.replace('-', '') + ' - AYAKTAN'
                    df_ayaktan['Value'] = df_ayaktan['Ayaktan Sayı']
                    all_qroups.append(df_ayaktan[['Həkim Adı', 'Ay', 'İl', 'Ay_Reqem', 'Qrup_Adı', 'Value']].copy())
                
                # 2. YATAN məlumatları olan sətirləri ayır (Yatan Sayı > 0 olan sətirlər)
                df_yatan = df_raw[df_raw['Yatan Sayı'] > 0].copy()
                if not df_yatan.empty:
                    df_yatan['Qrup_Adı'] = df_yatan['Məlumat Tipi'].str.replace('-isci', '').str.replace('-', '') + ' - YATAN'
                    df_yatan['Value'] = df_yatan['Yatan Sayı']
                    all_qroups.append(df_yatan[['Həkim Adı', 'Ay', 'İl', 'Ay_Reqem', 'Qrup_Adı', 'Value']].copy())
                
                # 3. Bütün Ayaktan və Yatan məlumatlarını birləşdir
                if not all_qroups:
                     raise ValueError("Seçilmiş həkimlər üçün heç bir fəaliyyət məlumatı tapılmadı (Ayaktan və Yatan yoxdur).")
                
                df_final = pd.concat(all_qroups, ignore_index=True)
                df_final['Ay_Adi'] = df_final['Ay'].apply(lambda x: datetime.strptime(str(x), '%Y-%m').strftime('%B %Y'))

                # Openpyxl ilə formatlaşdırma
                buffer = BytesIO()
                wb = Workbook()
                wb.remove(wb.active) # Default sheeti silirik

                thin_border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="31869B", end_color="31869B", fill_type="solid") # Tünd Mavi
                total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Açıq Yaşıl
                header_font = Font(bold=True, color="FFFFFF")
                title_font = Font(bold=True, size=14)
                
                # qrup_adi-ni for dövrəsindən kənarda təyin etmək üçün, dəyərini saxlayırıq (Əsasən Peşə Stat-da lazım olur)
                last_qrup_adi = "" 

                # Hər bir tip və xidmət növü üçün ayrı sheet yarat
                for qrup_adi, df_g in df_final.groupby('Qrup_Adı'):
                    last_qrup_adi = qrup_adi # Dəyəri saxla
                    
                    is_ayaktan = 'AYAKTAN' in qrup_adi
                    value_metric = 'Ayaktan Sayı' if is_ayaktan else 'Yatan Sayı'

                    pivot_df = df_g.pivot_table(index='Həkim Adı', columns='Ay_Adi', values='Value', aggfunc='sum', fill_value=0)
                    
                    if pivot_df.empty: continue

                    # Cəm sütunu əlavə et
                    pivot_df['CƏMİ'] = pivot_df.sum(axis=1)

                    # Cəm sətri əlavə et (TOPLAM)
                    total_row = pd.DataFrame(pivot_df.sum(axis=0)).T
                    total_row.index = ['TOPLAM']
                    
                    pivot_df = pd.concat([pivot_df, total_row], ignore_index=False)
                    
                    # Ayları düzgün sırayla almaq üçün sıralayırıq
                    aylar_sirasiz = df_g[['Ay_Adi', 'İl', 'Ay_Reqem']].drop_duplicates().sort_values(['İl', 'Ay_Reqem'])
                    aylar_sirasi = aylar_sirasiz['Ay_Adi'].tolist() + ['CƏMİ']
                    pivot_df = pivot_df.reindex(columns=aylar_sirasi)
                    
                    # Sheet adını yarat
                    sheet_name = qrup_adi.replace(' - ', '_').replace(' ', '_').replace('Gonderen_LAB', 'LAB_GOND')[:31]
                    ws = wb.create_sheet(sheet_name)
                    
                    # 1-ci sətir: Başlığı əlavə et və formatlaşdır
                    ws.cell(row=1, column=1, value=f"Statistika: {qrup_adi} ({start_date} - {end_date})")
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(aylar_sirasi) + 1)
                    ws['A1'].font = title_font
                    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 2-ci sətir: Sütun başlıqlarını yaz
                    ws.append(['Həkim Adı'] + list(pivot_df.columns))
                    
                    # 3-cü sətirdən: Dataları yaz
                    rows = dataframe_to_rows(pivot_df.reset_index().rename(columns={'index': 'Həkim Adı'}), header=False, index=False)
                    for r_idx, row in enumerate(rows, 3):
                        ws.append(row)

                    # Formatlaşdırma
                    # Sütun başlıqları üçün format (indiki 2-ci sətir)
                    for cell in ws[2]:
                        cell.border = thin_border
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # Məlumat və TOPLAM sətirləri üçün format
                    for row_cells in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        is_total_row = (row_cells[0].value == 'TOPLAM')
                        
                        for cell in row_cells:
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                            if is_total_row:
                                cell.fill = total_fill
                                cell.font = Font(bold=True)
                            
                            # Həkim Adı sütununu sola hizala
                            if cell.column == 1:
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                                
                    # Sütun enliklərini tənzimlə (Oxunaqlı olması üçün)
                    dims = {}
                    for row in ws.rows:
                        for cell in row:
                            if cell.value:
                                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                    for col, value in dims.items():
                        ws.column_dimensions[col].width = value + 3


                wb.save(buffer)
                buffer.seek(0)
                file_name = f"Statistika_detalli_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                await context.bot.send_document(
                    chat_id=update.effective_chat.id,
                    document=buffer,
                    filename=file_name,
                    caption=f"📊 {start_date} - {end_date} tarixləri üçün {category_name} statistikası (Detallı format)",
                    reply_to_message_id=update.effective_message.message_id
                )
                
                await update.effective_chat.send_message(
                    "✅ Excel faylı uğurla göndərildi və formatlaşdırıldı!\nNövbəti addım: Ana menyudan yeni bir seçim edə bilərsiniz.",
                    reply_markup=get_return_main_menu_button(),
                    reply_to_message_id=update.effective_message.message_id
                )
                return MAIN_MENU


    except Exception as e:
        logger.error(f"export_to_excel xətası: {e}")
        await update.effective_chat.send_message(
            f"Xəta baş verdi: {e}\nZəhmət olmasa, ana menyudan yeni bir seçim edin.",
            reply_markup=get_return_main_menu_button(),
            reply_to_message_id=update.effective_message.message_id
        )
        return MAIN_MENU

# Qrafik generasiya (Dəyişilməyib)
async def generate_graph(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("generate_graph funksiyası çağırıldı.")

    start_date = context.user_data.get('start_date', (datetime.now() - timedelta(days=180)).strftime('%d.%m.%Y'))
    end_date = context.user_data.get('end_date', datetime.now().strftime('%d.%m.%Y'))
    category = context.user_data.get('top10_category', 'hekim')
    specialty_name = context.user_data.get('specialty_name', 'Peşə')
    category_names = {
        'eden_ayaktan': 'Eden Ayaktan',
        'eden_yatan': 'Eden Yatan',
        'gonderen_ayaktan': 'Gonderen Ayaktan',
        'gonderen_yatan': 'Gonderen Yatan',
        'lab_gonderen_ayaktan': 'Lab Gonderen Ayaktan',
        'lab_gonderen_yatan': 'Lab Gonderen Yatan',
        'hekim': 'Həkim',
        'specialty': specialty_name
    }
    category_name = category_names[category]

    if not context.user_data.get('kodlar'):
        logger.warning("Həkim seçilməyib.")
        await update.effective_chat.send_message(
            "⚠️ Heç bir həkim seçilməyib. Əvvəlcə həkim əlavə edin və ya TOP 10 kateqoriyası seçin.",
            reply_markup=get_return_main_menu_button(),
            reply_to_message_id=update.effective_message.message_id
        )
        return MAIN_MENU

    try:
        logger.info(f"Generating graphs for category: {category_name}, doctors: {context.user_data['adlar']}")
        with oracledb.connect(user=username, password=password, dsn=dsn) as conn:
            category_config = {
                'eden_ayaktan': ('Eden-isci', 'YPERSID', query_template_specific, 'AYAKTAN'),
                'eden_yatan': ('Eden-isci', 'YPERSID', query_template_specific, 'YATAN'),
                'gonderen_ayaktan': ('Gonderen-isci', 'IPERSID', query_template_specific, 'AYAKTAN'),
                'gonderen_yatan': ('Gonderen-isci', 'IPERSID', query_template_specific, 'YATAN'),
                'lab_gonderen_ayaktan': ('Gonderen-LAB', 'IPERSID', query_template_lab_specific, 'AYAKTAN'),
                'lab_gonderen_yatan': ('Gonderen-LAB', 'IPERSID', query_template_lab_specific, 'YATAN'),
                'hekim': None,
                'specialty': None
            }

            if category in ['hekim', 'specialty']:
                query_sources = [
                    ("Eden-isci", "YPERSID", query_template_specific),
                    ("Gonderen-isci", "IPERSID", query_template_specific),
                    ("Gonderen-LAB", "IPERSID", query_template_lab_specific)
                ]
                combined_data = {}
                for tip, persid_col, _ in query_sources:
                    for col in ["AYAKTAN", "YATAN"]:
                        combined_data[(tip, col)] = pd.DataFrame()
            else:
                tip, persid_col, template, metric = category_config[category]
                query_sources = [(tip, persid_col, template)]
                combined_data = {(tip, metric): pd.DataFrame()}

            for kod, ad in zip(context.user_data['kodlar'], context.user_data['adlar']):
                for tip, persid_col, template in query_sources:
                    query = template.format(tip=tip, persid_col=persid_col)
                    logger.info(f"Executing query for doctor {ad} ({kod}), type {tip}")
                    df = pd.read_sql(query, con=conn, params={"kod": kod.strip(), "start_date": start_date, "end_date": end_date})
                    if df.empty:
                        logger.info(f"No data found for {ad} ({kod}) in {tip}")
                        continue
                    logger.info(f"Data retrieved for {ad} ({kod}) in {tip}: {df.shape[0]} rows")
                    for col in (["AYAKTAN", "YATAN"] if category in ['hekim', 'specialty'] else [category_config[category][3]]):
                        temp_df = df[["AY", col]].copy()
                        temp_df.rename(columns={col: ad}, inplace=True)
                        if combined_data.get((tip, col)) is not None:
                            if combined_data[(tip, col)].empty:
                                combined_data[(tip, col)] = temp_df
                            else:
                                combined_data[(tip, col)] = pd.merge(combined_data[(tip, col)], temp_df, on="AY", how="outer")

            if not any(df_plot.shape[0] > 0 for df_plot in combined_data.values()):
                logger.warning(f"{start_date} - {end_date} tarixləri üçün {category_name} qrafik məlumatı tapılmadı.")
                await update.effective_chat.send_message(
                    f"⚠️ Seçilmiş həkimlər üçün {start_date} - {end_date} tarixləri arasında {category_name} məlumat tapılmadı.",
                    reply_markup=get_return_main_menu_button(),
                    reply_to_message_id=update.effective_message.message_id
                )
                return MAIN_MENU

            for (tip, col), df_plot in combined_data.items():
                if not df_plot.empty:
                    logger.info(f"Generating graph for {tip} - {col}, rows: {df_plot.shape[0]}")
                    df_plot = df_plot.sort_values(by="AY")
                    plt.figure(figsize=(12, 6))
                    for idx, hekim in enumerate(df_plot.columns[1:]):
                        color = custom_colors[idx % len(custom_colors)]
                        plt.plot(df_plot["AY"], df_plot[hekim], marker='o', linewidth=2, label=f"{hekim} (cəmi: {int(df_plot[hekim].sum())})", color=color)
                        for x, y in zip(df_plot["AY"], df_plot[hekim]):
                            if pd.notnull(y):
                                plt.annotate(str(int(y)), (x, y), textcoords="offset points", xytext=(0, 8), ha='center',
                                             fontsize=8, color=color, weight='bold')
                    title = f"TOP 10 {category_name} – {col} Statistikası ({start_date} - {end_date})" if category not in ['hekim', 'specialty'] else f"{tip} – {col} Müqayisəli Qrafik ({start_date} - {end_date})"
                    plt.title(title)
                    plt.xlabel("Ay")
                    plt.ylabel("Say")
                    plt.xticks(rotation=45)
                    plt.legend()
                    plt.tight_layout()

                    buffer = BytesIO()
                    plt.savefig(buffer, format='png')
                    buffer.seek(0)
                    caption = f"TOP 10 {category_name} – {col} Statistikası" if category not in ['hekim', 'specialty'] else f"{tip} – {col} Statistikası"
                    await context.bot.send_photo(
                        chat_id=update.effective_chat.id,
                        photo=buffer,
                        caption=caption,
                        reply_to_message_id=update.effective_message.message_id
                    )
                    buffer.close()
                    plt.close()
                    logger.info(f"{tip} – {col} qrafiki göndərildi ({category_name}).")

            context.user_data['last_action'] = 'manual' if category == 'hekim' else 'top10'
            await update.effective_chat.send_message(
                f"📊 {category_name} statistikasını Excel olaraq göndərilsin mi?\n"
                "Hə: Excel faylı göndəriləcək.\nYox: Ana menyuya qayıdacaqsınız.",
                reply_markup=get_export_prompt_buttons(context.user_data['last_action']),
                reply_to_message_id=update.effective_message.message_id
            )
            return EXPORT_PROMPT

    except Exception as e:
        logger.error(f"generate_graph xətası ({category_name}): {e}")
        await update.effective_chat.send_message(
            f"Xəta baş verdi: {e}\nZəhmət olmasa, ana menyudan yeni bir seçim edin.",
            reply_markup=get_return_main_menu_button(),
            reply_to_message_id=update.effective_message.message_id
        )
        return MAIN_MENU

# Peşəyə görə qrafik generasiyası (Dəyişilməyib)
async def generate_specialty_graph(update: Update, context: ContextTypes.DEFAULT_TYPE, stat_type: str):
    logger.info(f"generate_specialty_graph funksiyası çağırıldı: {stat_type}")

    df = context.user_data.get('specialty_df')
    start_date = context.user_data.get('start_date')
    end_date = context.user_data.get('end_date')
    specialty_name = context.user_data.get('specialty_name', 'Peşə')

    stat_types = {
        'eden_ayaktan': ('Eden Ayaktan', 'EDEN_SAYISI', lambda x: x['HK_HASTATURU'].str.contains('Ayaktan', case=False, na=False) & (x['EDEN_SAYISI'] > 0)),
        'eden_yatan': ('Eden Yatan', 'EDEN_SAYISI', lambda x: x['HK_HASTATURU'].str.contains('Yatan', case=False, na=False) & (x['EDEN_SAYISI'] > 0)),
        'gonderen_ayaktan': ('Göndərən Ayaktan', 'GONDEREN_SAYISI', lambda x: x['HK_HASTATURU'].str.contains('Ayaktan', case=False, na=False) & (x['GONDEREN_SAYISI'] > 0)),
        'gonderen_yatan': ('Göndərən Yatan', 'GONDEREN_SAYISI', lambda x: x['HK_HASTATURU'].str.contains('Yatan', case=False, na=False) & (x['GONDEREN_SAYISI'] > 0)),
        'lab_gonderen_ayaktan': ('Laboratoriya Göndərən Ayaktan', 'GONDEREN_SAYISI', lambda x: x['HK_HASTATURU'].str.contains('Ayaktan', case=False, na=False) & x['ISLEMGRUPADI'].str.contains('lab', case=False, na=False) & (x['GONDEREN_SAYISI'] > 0)),
        'lab_gonderen_yatan': ('Laboratoriya Göndərən Yatan', 'GONDEREN_SAYISI', lambda x: x['HK_HASTATURU'].str.contains('Yatan', case=False, na=False) & x['ISLEMGRUPADI'].str.contains('lab', case=False, na=False) & (x['GONDEREN_SAYISI'] > 0))
    }

    if df is None or df.empty:
        logger.warning(f"{specialty_name} üçün qrafik məlumatı yoxdur.")
        await update.effective_chat.send_message(
            f"⚠️ {specialty_name} peşəsi üçün {start_date} - {end_date} tarixləri arasında məlumat tapılmadı.\n"
            "Ana menyudan yeni bir seçim edin.",
            reply_markup=get_return_main_menu_button(),
            reply_to_message_id=update.effective_message.message_id
        )
        return MAIN_MENU

    try:
        context.user_data['last_action'] = 'specialty' # Excel üçün

        if stat_type == 'specialty':  # Ümumi Statistika
            graphs_generated_count = 0
            for stat_key, (category_name, metric, filter_func) in stat_types.items():
                df_filtered = df[filter_func(df)].copy()
                if not df_filtered.empty:
                    await generate_single_graph(update, context, df_filtered, category_name, metric, start_date, end_date, specialty_name)
                    graphs_generated_count += 1
            
            if graphs_generated_count == 0:
                await update.effective_chat.send_message(
                    f"⚠️ {specialty_name} peşəsi üçün seçilmiş bütün alt kateqoriyalarda (Ümumi Statistika) məlumat tapılmadı.",
                    reply_to_message_id=update.effective_message.message_id
                )
        else:  # Specific statistic type
            category_name, metric, filter_func = stat_types[stat_type]
            df_filtered = df[filter_func(df)].copy()
            if df_filtered.empty:
                logger.warning(f"{specialty_name} üçün {category_name} məlumatı yoxdur.")
                await update.effective_chat.send_message(
                    f"⚠️ {specialty_name} peşəsi üçün {category_name} məlumat tapılmadı.",
                    reply_to_message_id=update.effective_message.message_id
                )
            else:
                await generate_single_graph(update, context, df_filtered, category_name, metric, start_date, end_date, specialty_name)

        # Excel ixrac sorğusu
        await prompt_excel_export(update, context, specialty_name)
        return EXPORT_PROMPT

    except Exception as e:
        logger.error(f"generate_specialty_graph xətası ({specialty_name}, {stat_type}): {e}")
        await update.effective_chat.send_message(
            f"Xəta baş verdi: {e}\nNövbəti addım: Başqa bir peşə seçin və ya ana menyuya qayıdın.",
            reply_markup=get_specialty_buttons(context.user_data.get('specialties', [])),
            reply_to_message_id=update.effective_message.message_id
        )
        return SELECT_SPECIALTY

# Helper function to generate individual graphs (Dəyişilməyib)
async def generate_single_graph(update: Update, context: ContextTypes.DEFAULT_TYPE, df_filtered, category_name, metric, start_date, end_date, specialty_name):
    df_filtered['TARIH'] = pd.to_datetime(df_filtered['AY_ADI'], format='%B %Y', errors='coerce', dayfirst=False)
    df_filtered = df_filtered.sort_values('TARIH')
    pivot_df = df_filtered.pivot_table(
        values=metric,
        index='AY_ADI',
        columns='DOKTOR_ADI',
        aggfunc='sum',
        fill_value=0
    )

    if not pivot_df.empty:
        pivot_df.index = pd.Categorical(
            pivot_df.index,
            categories=df_filtered['AY_ADI'].unique(),
            ordered=True
        )
        pivot_df = pivot_df.sort_index()

        plt.figure(figsize=(12, 6))
        for idx, hekim in enumerate(pivot_df.columns):
            color = custom_colors[idx % len(custom_colors)]
            plt.plot(pivot_df.index, pivot_df[hekim], marker='o', linewidth=2,
                     label=f"{hekim} (cəmi: {int(pivot_df[hekim].sum())})", color=color)
            for x, y in zip(pivot_df.index, pivot_df[hekim]):
                if y > 0:
                    plt.annotate(str(int(y)), (x, y), textcoords="offset points", xytext=(0, 8),
                                 ha='center', fontsize=8, color=color, weight='bold')

        title = f"{specialty_name} – {category_name} Statistikası ({start_date} - {end_date})"
        plt.title(title)
        plt.xlabel("Ay")
        plt.ylabel("Xidmət Sayı" if metric == 'EDEN_SAYISI' else "Göndərən Say")
        plt.xticks(rotation=45, ha='right')
        plt.legend()
        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        await context.bot.send_photo(
            chat_id=update.effective_chat.id,
            photo=buffer,
            caption=f"{specialty_name} – {category_name} Statistikası",
            reply_to_message_id=update.effective_message.message_id
        )
        buffer.close()
        plt.close()
        logger.info(f"{specialty_name} – {category_name} qrafiki göndərildi.")

# Helper function to prompt Excel export (Dəyişilməyib)
async def prompt_excel_export(update: Update, context: ContextTypes.DEFAULT_TYPE, specialty_name):
    await update.effective_chat.send_message(
        f"📊 {specialty_name} statistikasını Excel olaraq göndərilsin mi?\n"
        "Hə: Excel faylı göndəriləcək.\nYox: Ana menyuya qayıdacaqsınız.",
        reply_markup=get_export_prompt_buttons(context.user_data.get('last_action')),
        reply_to_message_id=update.effective_message.message_id
    )

# Peşəyə görə həkimləri seçmək (Dəyişilməyib)
async def select_specialty(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("select_specialty funksiyası çağırıldı.")

    try:
        with oracledb.connect(user=username, password=password, dsn=dsn) as conn:
            query = """
                SELECT PU_ID, PU_UNVAN
                FROM FONETHBYSADM.H_PUNVAN
                WHERE PU_ID IN (
                    90, 95, 96, 108, 114, 117, 3, 6, 7, 9, 17, 19, 22, 23, 25,
                    26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40,
                    41, 42, 43, 44, 45, 49, 60, 62, 63, 65, 66, 98, 118, 121, 125, 133, 127, 123, 126
                )
                ORDER BY PU_UNVAN
            """
            specialties_df = pd.read_sql(query, con=conn)
            specialties = specialties_df.to_dict('records')

            if not specialties:
                logger.warning("Heç bir peşə tapılmadı.")
                await update.effective_chat.send_message(
                    "⚠️ Heç bir peşə tapılmadı.\nNövbəti addım: Ana menyudan yeni bir seçim edin.",
                    reply_markup=get_return_main_menu_button(),
                    reply_to_message_id=update.effective_message.message_id
                )
                return MAIN_MENU

            await update.effective_chat.send_message(
                "👨‍⚕️ Peşəyə görə həkim statistikasını görmək üçün bir peşə seçin:\nNövbəti addım: Aşağıdakı peşələrdən birini seçin.",
                reply_markup=get_specialty_buttons(specialties),
                reply_to_message_id=update.effective_message.message_id
            )
            context.user_data['specialties'] = specialties
            logger.info(f"Peşələr göstərildi: {[s['PU_UNVAN'] for s in specialties]}")
            return SELECT_SPECIALTY

    except Exception as e:
        logger.error(f"select_specialty xətası: {e}")
        await update.effective_chat.send_message(
            f"Xəta baş verdi: {e}\nNövbəti addım: Ana menyudan yeni bir seçim edin.",
            reply_markup=get_return_main_menu_button(),
            reply_to_message_id=update.effective_message.message_id
        )
        return MAIN_MENU

# Düymələrə basıldıqda callback handler (DÜZƏLİŞLƏR)
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    logger.info(f"Button pressed: {query.data}, chat_id: {update.effective_chat.id}, message_id: {query.message.message_id}")
    await query.answer()

    # category_names lüğətini burada təyin edirik ki, lazım olan yerlərdə istifadə olunsun.
    category_names = {
        'eden_ayaktan': 'Eden Ayaktan', 'eden_yatan': 'Eden Yatan',
        'gonderen_ayaktan': 'Göndərən Ayaktan', 'gonderen_yatan': 'Göndərən Yatan',
        'lab_gonderen_ayaktan': 'Laboratoriya Göndərən Ayaktan', 'lab_gonderen_yatan': 'Laboratoriya Göndərən Yatan',
        'specialty': 'Ümumi Statistika'
    }

    top10_categories = {
        'menu_top10_eden_ayaktan': 'eden_ayaktan',
        'menu_top10_eden_yatan': 'eden_yatan',
        'menu_top10_gonderen_ayaktan': 'gonderen_ayaktan',
        'menu_top10_gonderen_yatan': 'gonderen_yatan',
        'menu_top10_lab_ayaktan': 'lab_gonderen_ayaktan',
        'menu_top10_lab_yatan': 'lab_gonderen_yatan'
    }

    stat_types = {
        'stat_eden_ayaktan': 'eden_ayaktan',
        'stat_eden_yatan': 'eden_yatan',
        'stat_gonderen_ayaktan': 'gonderen_ayaktan',
        'stat_gonderen_yatan': 'gonderen_yatan',
        'stat_lab_gonderen_ayaktan': 'lab_gonderen_ayaktan',
        'stat_lab_gonderen_yatan': 'lab_gonderen_yatan',
        'stat_umumi': 'specialty'
    }

    try:
        if query.data == 'menu_stats':
            # 1. Fərdi Həkim: İndi əvvəlcə Tarix Aralığı soruşulur.
            context.user_data.clear()
            context.user_data['kodlar'] = []
            context.user_data['adlar'] = []
            context.user_data['last_action'] = 'manual' # Fərdi axtarış rejimində olduğumuzu qeyd edirik
            
            await query.message.reply_text(
                "📅 Zəhmət olmasa, fərdi statistika üçün **tarix aralığını seçin**:\nNövbəti addım: 3 aylıq, 6 aylıq və ya 1 illik seçim edin.",
                reply_markup=get_date_range_buttons(),
                reply_to_message_id=query.message.message_id
            )
            return SELECT_DATE_RANGE
            
        elif query.data == 'menu_date':
            # 2. Tarixə Görə Analiz: İstifadəçini geri göndəririk.
            await query.message.reply_text(
                "⚠️ Tarix aralığını bu menyudan müstəqil seçə bilməzsiniz. Zəhmət olmasa, əvvəlcə **'TOP 10'**, **'Fərdi Statistikalar'** və ya **'Peşəyə Görə Həkimlər'** bölməsini seçin.",
                reply_markup=get_return_main_menu_button(),
                reply_to_message_id=query.message.message_id
            )
            return MAIN_MENU
            
        elif query.data in top10_categories:
            # 3. TOP 10 Seçimi: Peşə Seçiminə keçid.
            context.user_data.clear()
            context.user_data['kodlar'] = []
            context.user_data['adlar'] = []
            context.user_data['top10_category'] = top10_categories[query.data]
            context.user_data['last_action'] = 'top10'
            await query.message.reply_text(
                "👨‍⚕️ TOP 10 həkimlər üçün peşə seçin:\nNövbəti addım: Peşələrdən birini seçəcəksiniz.",
                reply_to_message_id=query.message.message_id
            )
            return await select_specialty(update, context)
            
        elif query.data == 'menu_specialty':
            # 4. Peşəyə Görə Həkimlər: Peşə Seçiminə keçid.
            context.user_data.clear()
            context.user_data['kodlar'] = []
            context.user_data['adlar'] = []
            context.user_data.pop('top10_category', None)
            context.user_data['last_action'] = 'specialty'
            
            await query.message.reply_text(
                "👨‍⚕️ Peşələr yüklənir, zəhmət olmasa gözləyin...\nNövbəti addım: Peşələrdən birini seçəcəksiniz.",
                reply_to_message_id=query.message.message_id
            )
            return await select_specialty(update, context)
            
        # --- Xüsusi düymələr ---
        elif query.data == 'menu_passive':
            # ❌ Passiv Həkimlər: Xarici linkə yönləndirir
            keyboard = [[InlineKeyboardButton("Yeni Bota Keçid", url=PASSIVE_DOCTOR_LINK)]]
            await query.message.reply_text(
                "❌ Passiv Həkimlər siyahısına baxmaq üçün digər botun olduğu qrupa yönləndirəm:",
                reply_markup=InlineKeyboardMarkup(keyboard),
                reply_to_message_id=query.message.message_id
            )
            return MAIN_MENU
            
        elif query.data == 'menu_excel':
            # Export üçün məlumat yoxlanılır
            if not context.user_data.get('kodlar') and not context.user_data.get('specialty_df'):
                await query.message.reply_text(
                    "⚠️ Əvvəlcə bir statistika hazırlayın (Fərdi, TOP 10 və ya Peşə), sonra Excel faylını ixrac edin.",
                    reply_markup=get_return_main_menu_button(),
                    reply_to_message_id=query.message.message_id
                )
                return MAIN_MENU
            
            await query.message.reply_text(
                "📊 Excel faylı hazırlanır, zəhmət olmasa gözləyin...",
                reply_to_message_id=query.message.message_id
            )
            await export_to_excel(update, context)
            return MAIN_MENU
            
        elif query.data == 'menu_reset' or query.data == 'return_main_menu':
            # 🔁 Yeni Seçim / Ana menyuya qayıt
            context.user_data.clear()
            context.user_data['kodlar'] = []
            context.user_data['adlar'] = []
            await query.message.reply_text(
                "✅ Ana menyuya qayıdırsınız.\nNövbəti addım: Aşağıdakı menyudan bir seçim edin.",
                reply_markup=get_main_menu(),
                reply_to_message_id=query.message.message_id
            )
            return MAIN_MENU
        
        elif query.data.startswith('specialty_'):
            # 5. Peşə Seçimi (TOP 10 və ya Peşə Stat): Tarix Aralığına Keçid.
            pu_id = query.data[len('specialty_'):]
            specialty_name = next((s['PU_UNVAN'] for s in context.user_data.get('specialties', []) if str(s['PU_ID']) == pu_id), pu_id)
            context.user_data['specialty_id'] = pu_id
            context.user_data['specialty_name'] = specialty_name
            action_type = "TOP 10" if context.user_data.get('last_action') == 'top10' else "statistika"

            await query.message.reply_text(
                f"📅 {specialty_name} peşəsi üçün {action_type} **tarix aralığını seçin**:\nNövbəti addım: 3 aylıq, 6 aylıq və ya 1 illik seçim edin.",
                reply_markup=get_date_range_buttons(),
                reply_to_message_id=query.message.message_id
            )
            return SELECT_DATE_RANGE
        
        elif query.data in ['date_3m', 'date_6m', 'date_1y']:
            # 6. Tarix Aralığı Seçimi: Nəticəyə/Növbəti Addıma Keçid
            days = {'date_3m': 90, 'date_6m': 180, 'date_1y': 365}
            context.user_data['date_range'] = days[query.data]
            end = datetime.now()
            start = end - timedelta(days=days[query.data])
            context.user_data['start_date'] = start.strftime('%d.%m.%Y')
            context.user_data['end_date'] = end.strftime('%d.%m.%Y')
            logger.info(f"Tarix aralığı seçildi: {days[query.data]} gün. Action: {context.user_data.get('last_action')}")

            if context.user_data.get('last_action') == 'manual' or context.user_data.get('last_action') == 'manual_reselect_date':
                # Fərdi (manual) axtarış
                
                if context.user_data.get('last_action') == 'manual_reselect_date' and context.user_data.get('kodlar'):
                    # Təkrar tarix seçimi sonrası birbaşa qrafik hazırlama
                    await query.message.reply_text(
                        "📊 Yeni tarix aralığı üçün statistikalar hazırlanır, zəhmət olmasa gözləyin...",
                        reply_to_message_id=query.message.message_id
                    )
                    context.user_data['last_action'] = 'manual' # Normal manual rejimə qayıdırıq
                    await generate_graph(update, context) 
                    return EXPORT_PROMPT
                
                else: # Bu, menu_stats'dan sonra ilk tarix seçimidir.
                    await query.message.reply_text(
                        f"🔍 Həkim statistikası (Cari tarix aralığı: {context.user_data['start_date']} - {context.user_data['end_date']}) üçün **həkimin adını, soyadını və ya P_KODU-nu göndərin**:\n"
                        "Növbəti addım: Həkim məlumatlarını daxil edin.",
                        reply_to_message_id=query.message.message_id
                    )
                    return ASK_DOCTOR

            elif context.user_data.get('last_action') == 'top10':
                # TOP 10 funksiyasını çağır (Peşə və Tarix artıq seçilib)
                category_name = top10_categories.get(context.user_data.get('top10_category'), 'Həkim')
                await query.message.reply_text(
                    f"🏆 TOP 10 {category_name} Həkim siyahısı hazırlanır, zəhmət olmasa gözləyin...",
                    reply_to_message_id=query.message.message_id
                )
                await top_10_doctors(update, context)
                return EXPORT_PROMPT
            
            elif context.user_data.get('last_action') == 'specialty':
                # Peşə statistikasının növünü soruş
                await query.message.reply_text(
                    f"📊 {context.user_data.get('specialty_name', 'Peşə')} statistikası üçün hansı statistik növünü seçmək istəyirsiniz?",
                    reply_markup=get_stat_type_buttons(),
                    reply_to_message_id=query.message.message_id
                )
                return SELECT_STAT_TYPE
            
            else:
                await query.message.reply_text(
                    "⚠️ Əvvəlcə bir kateqoriya (TOP 10, Fərdi və ya Peşə) seçin.",
                    reply_markup=get_return_main_menu_button(),
                    reply_to_message_id=query.message.message_id
                )
                return MAIN_MENU
        
        # --- Fərdi Axtarışdan Sonra Tarix Dəyişikliyi ---
        elif query.data == 'back_to_date_range_manual':
            # Fərdi (manual) axtarışdan sonra tarix dəyişməyə qayıt
            if context.user_data.get('last_action') == 'manual' and context.user_data.get('kodlar'):
                context.user_data['last_action'] = 'manual_reselect_date' # Yeni vəziyyət qeyd edilir
                
                await query.message.reply_text(
                    f"📅 Seçilmiş həkimlər ({', '.join(context.user_data['adlar'])}) üçün yeni tarix aralığını seçin:",
                    reply_markup=get_date_range_buttons(),
                    reply_to_message_id=query.message.message_id
                )
                return SELECT_DATE_RANGE
            else:
                await query.message.reply_text(
                    "⚠️ Tarix aralığını dəyişmək üçün əvvəlcə Fərdi Statistikalar seçimi ilə ən az bir həkim əlavə edin.",
                    reply_markup=get_return_main_menu_button(),
                    reply_to_message_id=query.message.message_id
                )
                return MAIN_MENU

        # --- Digər Mərhələlər (Dəyişilməyib) ---
        elif query.data == 'continue_yes':
            await query.message.reply_text(
                "🔍 Yeni həkim axtarışı üçün ad, soyad və ya P_KODU daxil edin:\nNövbəti addım: Həkim məlumatlarını daxil edin.",
                reply_to_message_id=query.message.message_id
            )
            return ASK_DOCTOR
        elif query.data == 'continue_no':
            if not context.user_data.get('kodlar'):
                await query.message.reply_text(
                    "⚠️ Heç bir həkim seçilməyib. Əvvəlcə həkim əlavə edin.",
                    reply_to_message_id=query.message.message_id
                )
                return ASK_DOCTOR
            await query.message.reply_text(
                "📊 Statistikalar hazırlanır, zəhmət olmasa gözləyin...",
                reply_to_message_id=query.message.message_id
            )
            await generate_graph(update, context) # generate_graph EXPORT_PROMPT-ə keçir
            return EXPORT_PROMPT
        elif query.data == 'export_yes':
            await query.message.reply_text(
                "📊 Excel faylı hazırlanır, zəhmət olmasa gözləyin...",
                reply_to_message_id=query.message.message_id
            )
            await export_to_excel(update, context)
            return MAIN_MENU
        
        # Peşə statistikası geri qayıt düymələri
        elif query.data == 'back_to_date_range':
            context.user_data.pop('start_date', None)
            context.user_data.pop('end_date', None)
            context.user_data.pop('date_range', None)
            action_type = "TOP 10" if context.user_data.get('last_action') == 'top10' else "statistika"
            await query.message.reply_text(
                f"📅 {context.user_data.get('specialty_name', 'Peşə')} peşəsi üçün {action_type} tarix aralığını seçin:\nNövbəti addım: 3 aylıq, 6 aylıq və ya 1 illik seçim edin.",
                reply_markup=get_date_range_buttons(),
                reply_to_message_id=query.message.message_id
            )
            return SELECT_DATE_RANGE
        elif query.data == 'back_to_specialty':
            context.user_data.pop('specialty_id', None)
            context.user_data.pop('specialty_name', None)
            context.user_data.pop('start_date', None)
            context.user_data.pop('end_date', None)
            context.user_data.pop('date_range', None)
            await query.message.reply_text(
                "👨‍⚕️ Peşələr yüklənir, zəhmət olmasa gözləyin...\nNövbəti addım: Peşələrdən birini seçəcəksiniz.",
                reply_to_message_id=query.message.message_id
            )
            return await select_specialty(update, context)

        elif query.data.startswith('select_doctor_'):
            # Həkim seçimi tamamlananda (tək və ya çoxlu axtarışdan sonra)
            kod = query.data[len('select_doctor_'):]
            df = context.user_data.get('search_list')
            hekim_adi = df[df['P_KODU'] == kod]['HEKIM_ADI'].values[0]
            context.user_data['kodlar'].append(kod)
            context.user_data['adlar'].append(hekim_adi)
            del context.user_data['search_list']
            await query.message.reply_text(f"✅ {hekim_adi} əlavə edildi.")
            await query.message.reply_text(
                "Başqa həkim əlavə etmək istəyirsiniz?\nNövbəti addım: Hə seçsəniz yeni həkim əlavə edə, Yox seçsəniz statistikaları görə bilərsiniz.",
                reply_markup=get_continue_buttons(context.user_data['last_action']),
                reply_to_message_id=query.message.message_id
            )
            return CONTINUE_DOCTOR

        elif query.data in stat_types:
            # 7. Peşə Statistikasının Növü Seçimi: Qrafik və Nəticə
            stat_type = stat_types[query.data]
            
            try:
                with oracledb.connect(user=username, password=password, dsn=dsn) as conn:
                    if 'start_date' not in context.user_data or 'specialty_id' not in context.user_data:
                        raise ValueError("Tarix aralığı və ya peşə seçilməyib.")

                    df = pd.read_sql(query_template_specialty, con=conn, params={
                        "pu_id": int(context.user_data['specialty_id']),
                        "start_date": context.user_data['start_date'],
                        "end_date": context.user_data['end_date']
                    })

                    if df.empty:
                        raise ValueError("Məlumat tapılmadı.")

                    context.user_data['specialty_df'] = df
                    
                    await query.message.reply_text(
                        f"📊 {context.user_data['specialty_name']} peşəsi üçün {category_names[stat_type]} statistikaları hazırlanır, zəhmət olmasa gözləyin...",
                        reply_to_message_id=query.message.message_id
                    )
                    await generate_specialty_graph(update, context, stat_type)
                    return EXPORT_PROMPT

            except ValueError as ve:
                await query.message.reply_text(
                    f"⚠️ {context.user_data.get('specialty_name', 'Peşə')} üçün məlumat tapılmadı: {category_names.get(stat_type)}.\nNövbəti addım: Başqa bir statistik növü seçin.",
                    reply_markup=get_stat_type_buttons(),
                    reply_to_message_id=query.message.message_id
                )
                return SELECT_STAT_TYPE
            except Exception as e:
                logger.error(f"Peşə statistikası alınarkən ümumi xəta: {e}")
                await query.message.reply_text(
                    f"Xəta baş verdi: {e}\nNövbəti addım: Başqa bir peşə seçin və ya ana menyuya qayıdın.",
                    reply_markup=get_specialty_buttons(context.user_data.get('specialties', [])),
                    reply_to_message_id=query.message.message_id
                )
                return SELECT_SPECIALTY


    except Exception as e:
        logger.error(f"button_handler xətası: {e}")
        await query.message.reply_text(
            f"Xəta baş verdi: {e}\nZəhmət olmasa, ana menyudan yeni bir seçim edin.",
            reply_markup=get_return_main_menu_button(),
            reply_to_message_id=query.message.message_id
        )
        return MAIN_MENU

# Həkim sorğusu
async def ask_doctor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    search_input = text
    
    start_date = context.user_data.get('start_date')
    end_date = context.user_data.get('end_date')
    days = context.user_data.get('date_range', 365) 

    logger.info(f"Həkim axtarışı: {search_input}, Tarix aralığı: {start_date} - {end_date}, chat_id: {update.effective_chat.id}")

    await update.effective_chat.send_message(
        f"📅 Axtarış tarix aralığı: {start_date} - {end_date} ({days} gün)\n"
        "Növbəti addım: Həkim tapıldıqdan sonra daha çox həkim əlavə edə və ya statistikaları görə bilərsiniz.",
        reply_to_message_id=update.message.message_id
    )

    try:
        with oracledb.connect(user=username, password=password, dsn=dsn) as conn:
            query_list = f"""
                SELECT P_KODU, P_AD || ' ' || P_SOYAD AS HEKIM_ADI
                FROM FONETHBYSADM.H_PERSON
                WHERE LOWER(P_AD || ' ' || P_SOYAD) LIKE LOWER('%{search_input}%') OR LOWER(P_KODU) = LOWER('{search_input}')
            """
            df = pd.read_sql(query_list, con=conn)

            if df.empty:
                await update.effective_chat.send_message(
                    "Heç bir uyğun həkim tapılmadı. Yenidən ad, soyad və ya P_KODU göndərin:\nNövbəti addım: Doğru həkim məlumatlarını daxil edin.",
                    reply_to_message_id=update.message.message_id
                )
                logger.warning(f"Həkim tapılmadı: {search_input}")
                return ASK_DOCTOR

            if len(df) > 1:
                keyboard = []
                for _, row in df.iterrows():
                    button_text = f"{row['HEKIM_ADI']} ({row['P_KODU']})"
                    callback_data = f"select_doctor_{row['P_KODU']}"
                    keyboard.append([InlineKeyboardButton(button_text, callback_data=callback_data)])
                keyboard.append([InlineKeyboardButton("Ana menyuya qayıt", callback_data='return_main_menu')])
                reply_markup = InlineKeyboardMarkup(keyboard)
                await update.effective_chat.send_message(
                    "Bir neçə uyğun həkim tapıldı. Zəhmət olmasa birini seçin:",
                    reply_markup=reply_markup,
                    reply_to_message_id=update.message.message_id
                )
                context.user_data['search_list'] = df
                logger.info(f"Çoxsaylı həkim tapıldı: {len(df)}")
                return ASK_DOCTOR

            if len(df) == 1:
                kod = df.iloc[0]['P_KODU']
                hekim_adi = df.iloc[0]['HEKIM_ADI']
                context.user_data['kodlar'].append(kod)
                context.user_data['adlar'].append(hekim_adi)
                await update.effective_chat.send_message(f"✅ {hekim_adi} əlavə edildi.")
                await update.effective_chat.send_message(
                    "Başqa həkim əlavə etmək istəyirsiniz?\nNövbəti addım: Hə seçsəniz yeni həkim əlavə edə, Yox seçsəniz statistikaları görə bilərsiniz.",
                    reply_markup=get_continue_buttons(context.user_data['last_action']),
                    reply_to_message_id=update.message.message_id
                )
                logger.info(f"Həkim əlavə edildi: {hekim_adi}")
                return CONTINUE_DOCTOR
                
    except Exception as e:
        logger.error(f"ask_doctor xətası: {e}")
        await update.effective_chat.send_message(
            f"Xəta baş verdi: {e}\nNövbəti addım: Yenidən cəhd edin və ya ana menyudan davam edin.",
            reply_markup=get_return_main_menu_button(),
            reply_to_message_id=update.message.message_id
        )
        return ASK_DOCTOR

# Əsas proqram
if __name__ == '__main__':
    app = ApplicationBuilder().token(bot_token).build()

    # /start komandasını işə salır
    app.add_handler(CommandHandler("start", start))
    
    # Bütün düymə kliklərini idarə edir
    app.add_handler(CallbackQueryHandler(button_handler))
    
    # ASK_DOCTOR mərhələsində daxil edilən mətn mesajlarını idarə edir.
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, ask_doctor))

    # Botu işə sal
    logger.info("Bot işə salındı...")
    app.run_polling(poll_interval=1.0)