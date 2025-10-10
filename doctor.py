import oracledb
import pandas as pd
import os
import logging
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
import sys
import traceback
import re
import argparse
from datetime import datetime

# Loglama ayarlarını yapılandırır
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Konfigürasyon: Varsayılan değerler
DEFAULT_ORACLE_CLIENT_PATH = os.getenv('ORACLE_CLIENT_PATH', r"C:\instant\instantclient_23_9")
DEFAULT_OUTPUT_PATH = os.getenv('OUTPUT_PATH', os.path.join(os.path.expanduser('~'), 'Desktop', 'raport.xlsx'))
DEFAULT_USERNAME = os.getenv('DB_USERNAME', "NURAN")
DEFAULT_PASSWORD = os.getenv('DB_PASSWORD', "Nuran..2024!!")
DEFAULT_DSN = os.getenv('DB_DSN', "172.18.79.23:1521/FONETAZ")
DEFAULT_START_DATE = "01.01.2025"
DEFAULT_END_DATE = datetime.now().strftime("%d.%m.%Y")

# SQL sorgusu
SQL_QUERY = """
WITH filtered_data AS (
    SELECT
        NVL(t.FirmaAdi, 'Bilinmeyen Firma') AS FirmaAdi,
        t.FISKODU,
        t.SNLMALZEMEADI,
        t.LOT_NO,
        t.BIRIMADI,
        t.TARIH,
        t.FIYATKDV,
        SUM(t.MIKTAR) AS MIKTAR,
        SUM(t.TOPLAMCIKIS) AS TOPLAMCIKIS,
        SUM(t.TUTAR) AS TUTAR,
        SUM(t.MIKTAR * t.BIRIMCARPAN - t.TOPLAMCIKIS) AS GIRISKALAN
    FROM fonethbys.V_STOK_LIST_GIRISFISDETAY t
    WHERE t.FORMTYPE IN (1, 6)
        AND t.TARIH >= TO_DATE(:start_date, 'DD.MM.YYYY')
        AND t.TARIH < TO_DATE(:end_date, 'DD.MM.YYYY')
        AND t.DEPOID = 2
    GROUP BY
        NVL(t.FirmaAdi, 'Bilinmeyen Firma'),
        t.FISKODU,
        t.SNLMALZEMEADI,
        t.LOT_NO,
        t.BIRIMADI,
        t.TARIH,
        t.FIYATKDV
),
totals AS (
    SELECT
        FirmaAdi,
        NULL AS FISKODU,
        NULL AS SNLMALZEMEADI,
        NULL AS LOT_NO,
        NULL AS BIRIMADI,
        NULL AS TARIH,
        NULL AS FIYATKDV,
        SUM(MIKTAR) AS MIKTAR,
        SUM(TOPLAMCIKIS) AS TOPLAMCIKIS,
        SUM(TUTAR) AS TUTAR,
        SUM(GIRISKALAN) AS GIRISKALAN
    FROM filtered_data
    GROUP BY FirmaAdi
)
SELECT
    ROW_NUMBER() OVER (PARTITION BY FirmaAdi ORDER BY TARIH, FISKODU) AS "No",
    FISKODU,
    SNLMALZEMEADI,
    LOT_NO,
    BIRIMADI,
    TARIH,
    FIYATKDV,
    MIKTAR,
    TOPLAMCIKIS,
    TUTAR,
    GIRISKALAN,
    FirmaAdi
FROM filtered_data
UNION ALL
SELECT
    NULL AS "No",
    FISKODU,
    SNLMALZEMEADI,
    LOT_NO,
    BIRIMADI,
    TARIH,
    FIYATKDV,
    MIKTAR,
    TOPLAMCIKIS,
    TUTAR,
    GIRISKALAN,
    FirmaAdi
FROM totals
ORDER BY FirmaAdi, "No" NULLS LAST
"""

def initialize_oracle_client(client_path):
    """Oracle Instant Client'ı başlatır."""
    try:
        if not os.path.exists(client_path):
            logger.error(f"Oracle Instant Client yolu bulunamadı: {client_path}")
            sys.exit(1)
        oracledb.init_oracle_client(lib_dir=client_path)
        logger.info(f"Oracle Instant Client başlatıldı: {client_path}")
    except oracledb.DatabaseError as e:
        logger.error(f"Oracle Client başlatılamadı: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Beklenmeyen hata Oracle Client başlatılırken: {e}")
        sys.exit(1)

def check_file_access(path):
    """Çıktı dosyasının yazılabilir olduğunu kontrol eder."""
    try:
        directory = os.path.dirname(path)
        if not os.path.exists(directory):
            os.makedirs(directory)
            logger.info(f"Çıktı dizini oluşturuldu: {directory}")
        if not os.access(directory, os.W_OK):
            logger.error(f"Yazma izni yok: {directory}")
            sys.exit(1)
    except PermissionError:
        logger.error(f"❌ Dosyaya yazma izni yok: {path}. Lütfen dosyayı kapatın veya izinleri kontrol edin.")
        sys.exit(1)

def get_unique_output_path(base_path, overwrite=False):
    """Eğer dosya varsa ve overwrite False ise, benzersiz bir dosya adı oluşturur."""
    if overwrite or not os.path.exists(base_path):
        return base_path
    base, ext = os.path.splitext(base_path)
    counter = 1
    while True:
        new_path = f"{base}_{counter}{ext}"
        if not os.path.exists(new_path):
            return new_path
        counter += 1

def sanitize_sheet_name(name):
    """Excel sheet adını 31 karaktere kısaltır ve geçersiz karakterleri kaldırır."""
    name = re.sub(r'[\[\]\*?:/\\]', '_', name or "Sheet")
    return name[:31]

def generate_report(username, password, dsn, oracle_client_path, output_path, start_date, end_date, overwrite=False):
    """Veritabanından rapor oluşturur ve her firma için ayrı sheet'e kaydeder."""
    try:
        # Oracle Instant Client'ı başlat
        initialize_oracle_client(oracle_client_path)

        # Benzersiz dosya yolu belirle
        output_path = get_unique_output_path(output_path, overwrite)

        # SQLAlchemy engine oluştur
        logger.info("Veritabanına bağlanılıyor...")
        engine = create_engine(f'oracle+oracledb://{username}:{password}@{dsn}')

        # SQL sorgusunu çalıştır ve DataFrame'e yükle
        logger.info("SQL sorgusu çalıştırılıyor...")
        with engine.connect() as connection:
            df = pd.read_sql(SQL_QUERY, con=connection, params={'start_date': start_date, 'end_date': end_date})

        # DataFrame boş mu kontrol et
        if df.empty:
            logger.error("❌ Sorgudan veri alınamadı. View veya veriler kontrol edilmeli.")
            print("❌ Sorgudan veri alınamadı. View veya veriler kontrol edilmeli.")
            sys.exit(1)

        # DataFrame sütunlarını yazdır
        logger.info(f"DataFrame Columns: {df.columns.tolist()}")
        logger.info(f"Toplam {len(df)} satır veri alındı.")

        # FirmaAdi sütununu bul
        firma_col = None
        for col in df.columns:
            if col.lower() == 'firmaadi':
                firma_col = col
                break
        if not firma_col:
            logger.error("❌ 'FirmaAdi' sütunu DataFrame'de bulunamadı.")
            print("❌ 'FirmaAdi' sütunu DataFrame'de bulunamadı.")
            sys.exit(1)

        # Sütun adlarını eşleştirme
        column_mapping = {
            'No': 'No',
            'fiskodu': 'Fiş Kodu',
            'snlmalzemeadi': 'Malzeme Adı',
            'lot_no': 'Lot No',
            'birimadi': 'Birim',
            'tarih': 'Tarix',
            'fiyatkdv': 'Fiyat (KDV)',
            'miktar': 'Miqdar',
            'toplamcikis': 'Toplam Çıxış',
            'tutar': 'Tutar',
            'giriskalan': 'Giriş Qalan'
        }

        # DataFrame sütunlarını yeniden adlandır
        df = df.rename(columns=column_mapping)

        # Excel dosyası oluştur
        logger.info(f"Excel dosyası oluşturuluyor: {output_path}")
        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

        # Stil tanımlamaları
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        header_font = Font(bold=True)
        data_font = Font(bold=False)
        total_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        no_border = Border()

        # Sütun başlıkları (yerelleştirilmiş)
        headers = ["No", "Fiş Kodu", "Malzeme Adı", "Lot No", "Birim", "Tarix", "Fiyat (KDV)", "Miqdar", "Toplam Çıxış", "Tutar", "Giriş Qalan"]

        # Firmalara göre veriyi gruplandır
        grouped = df.groupby(firma_col)
        for firma_adi, group_df in grouped:
            # Sheet adını hazırla
            sheet_name = sanitize_sheet_name(firma_adi)
            worksheet = workbook.create_sheet(title=sheet_name)

            # İndeksi sıfırla
            group_df = group_df.reset_index(drop=True)

            # Seçili sütunları al
            group_df = group_df[headers]

            # Başlık satırı ekle
            title_cell = worksheet.cell(row=1, column=1, value=f"{firma_adi} üzrə Stok Giriş Hesabatı")
            title_cell.alignment = center_alignment
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            worksheet.merge_cells(f"A1:{get_column_letter(len(headers))}1")

            # Sütun başlıklarını yaz
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=2, column=col, value=header)
                cell.alignment = center_alignment
                cell.font = header_font
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                cell.border = thin_border

            # Verileri yaz ve kenarlık uygula
            for row_idx, row in enumerate(group_df.itertuples(index=False), start=0):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx + 3, column=col_idx, value=value)
                    cell.border = thin_border
                    if col_idx == 1 and value is None:  # Totals satırı
                        cell.font = total_font
                        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    else:
                        cell.font = data_font
                    if col_idx in (2, 3, 4, 5):  # Sola hizalı sütunlar
                        cell.alignment = left_alignment
                    elif col_idx == 1:  # Ortalı sütun
                        cell.alignment = center_alignment
                    else:  # Sağa hizalı sütunlar
                        cell.alignment = right_alignment
                        # Sayısal sütunlar için format
                        if col_idx in (7, 8, 9, 10, 11) and value is not None:
                            cell.number_format = '#,##0.00'

            # Boş satır ekle
            if None in group_df['No'].values:
                last_row = len(group_df) + 3
                for col in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=last_row + 1, column=col)
                    cell.border = no_border

            # Sütun genişliklerini ayarla
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = get_column_letter(col)
                for row in worksheet[column]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                if col == 3:  # Malzeme Adı
                    adjusted_width = max(adjusted_width, 30)
                elif col in (2, 4, 5):  # Fiş Kodu, Lot No, Birim
                    adjusted_width = max(adjusted_width, 15)
                elif col == 1:  # No
                    adjusted_width = max(adjusted_width, 10)
                else:  # Sayısal sütunlar
                    adjusted_width = max(adjusted_width, 12)
                worksheet.column_dimensions[column].width = min(adjusted_width, 40)

            # İlk iki satırı dondur
            worksheet.freeze_panes = 'A3'

            logger.info(f"Sheet oluşturuldu: {sheet_name} ({len(group_df)} satır)")

        # Excel dosyasını kaydet
        workbook.save(output_path)
        logger.info(f"✅ Rapor başarıyla '{output_path}' dosyasına kaydedildi.")
        print(f"✅ Rapor başarıyla '{output_path}' dosyasına kaydedildi.")

    except oracledb.DatabaseError as e:
        logger.error(f"❌ Veritabanı hatası: {e}\n{traceback.format_exc()}")
        print(f"❌ Veritabanı hatası: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"❌ Hata: {e}\n{traceback.format_exc()}")
        print(f"❌ Hata: {e}")
        sys.exit(1)

def main():
    # Komut satırı argümanlarını tanımla
    parser = argparse.ArgumentParser(description="Stok Giriş Raporu Oluşturucu")
    parser.add_argument('--username', default=DEFAULT_USERNAME, help='Veritabanı kullanıcı adı')
    parser.add_argument('--password', default=DEFAULT_PASSWORD, help='Veritabanı şifresi')
    parser.add_argument('--dsn', default=DEFAULT_DSN, help='Veritabanı DSN')
    parser.add_argument('--oracle-client-path', default=DEFAULT_ORACLE_CLIENT_PATH, help='Oracle Instant Client yolu')
    parser.add_argument('--output-path', default=DEFAULT_OUTPUT_PATH, help='Çıktı Excel dosya yolu')
    parser.add_argument('--start-date', default=DEFAULT_START_DATE, help='Başlangıç tarihi (DD.MM.YYYY)')
    parser.add_argument('--end-date', default=DEFAULT_END_DATE, help='Bitiş tarihi (DD.MM.YYYY)')
    parser.add_argument('--overwrite', action='store_true', help='Var olan dosyayı üzerine yaz')

    args = parser.parse_args()

    # Bağımlılıkları kontrol et
    required_modules = ['oracledb', 'pandas', 'openpyxl', 'sqlalchemy']
    for module in required_modules:
        try:
            __import__(module)
        except ImportError:
            logger.error(f"{module} modülü yüklü değil. Lütfen yükleyin: pip install {module}")
            sys.exit(1)

    # Raporu oluştur
    generate_report(
        username=args.username,
        password=args.password,
        dsn=args.dsn,
        oracle_client_path=args.oracle_client_path,
        output_path=args.output_path,
        start_date=args.start_date,
        end_date=args.end_date,
        overwrite=args.overwrite
    )

if __name__ == '__main__':
    main()